"""Batch recipe/workflow/HAL generation API routes."""
import asyncio
import csv
import io
from datetime import datetime
from pathlib import Path
from typing import Callable, List

from fastapi import APIRouter, Request


def create_batch_router(state, LIBRARY_DIR: Path, plate_tracking: dict, save_plate_tracking: Callable, plate_tracking_lock: asyncio.Lock) -> APIRouter:
    router = APIRouter(prefix="/api/batch", tags=["batch"])

    def generate_pipetting_recipe(plates: List[dict], hal_config: str, prefix: str, timestamp: str) -> dict:
        """Generate a JSON recipe for pipetting operations."""
        all_operations = []
        plate_ids = []
        for plate in plates:
            plate_ids.append(plate["plate_id"])
            for op in plate["operations"]:
                all_operations.append(op)

        operations_by_liquid: dict = {}
        for op in all_operations:
            liquid = op.get("liquid") or "default"
            if liquid not in operations_by_liquid:
                operations_by_liquid[liquid] = []
            operations_by_liquid[liquid].append(op)

        steps = []
        for liquid, ops in operations_by_liquid.items():
            transfers = [
                {
                    "dest_well": op["well"],
                    "volume_ul": op["volume"],
                    "source_well": op.get("source_well") or "A1",
                    "plate_id": op.get("plate_id")
                }
                for op in ops
            ]
            steps.append({"command": "TransferLiquid", "liquid": liquid, "transfers": transfers})

        return {
            "name": f"Excel_Pipetting_{prefix}_{timestamp}",
            "description": f"Auto-generated from Excel. Plates: {', '.join(plate_ids)}",
            "plate_ids": plate_ids,
            "hal_config": hal_config,
            "steps": steps,
            "metadata": {
                "generated": datetime.now().isoformat(),
                "total_operations": len(all_operations),
                "source": "excel_import"
            }
        }

    @router.post("/parse-excel")
    async def parse_excel_pipetting_plan(request: Request):
        """
        Parse a CSV/Excel pipetting plan.

        Accepts FormData (file upload) or JSON ({library_file}).
        Returns recipes split by plate capacity with plate IDs for traceability.
        """
        try:
            content_type = request.headers.get("content-type", "")

            if "application/json" in content_type:
                data = await request.json()
                library_file = data.get("library_file")
                plate_type = int(data.get("plate_type", 96))
                overflow_strategy = data.get("overflow_strategy", "multi_recipe")
                plate_prefix = data.get("plate_prefix", "PLATE")
                hal_config = data.get("hal_config", "")

                if not library_file:
                    return {"error": "No library file specified"}
                file_path = LIBRARY_DIR / "PipettingPlans" / library_file
                if not file_path.exists():
                    return {"error": f"File not found: {library_file}"}
                with open(file_path, "rb") as f:
                    content = f.read()
                filename = library_file.lower()
                source_file_name = library_file
            else:
                form = await request.form()
                file = form.get("file")
                plate_type = int(str(form.get("plate_type", 96)))
                overflow_strategy = str(form.get("overflow_strategy", "multi_recipe"))
                plate_prefix = str(form.get("plate_prefix", "PLATE"))
                hal_config = str(form.get("hal_config", ""))

                if not file:
                    return {"error": "No file uploaded"}
                if not file.filename:
                    return {"error": "File has no filename"}
                content = await file.read()
                filename = file.filename.lower()
                source_file_name = file.filename

            operations: list = []

            if filename.endswith(".csv"):
                text = content.decode("utf-8")
                reader = csv.DictReader(io.StringIO(text))
                for row in reader:
                    well = row.get("Well") or row.get("well") or row.get("Destination") or row.get("dest") or ""
                    liquid = row.get("Liquid") or row.get("Source") or row.get("source") or row.get("liquid") or ""
                    volume = row.get("Volume") or row.get("volume") or row.get("Volume_uL") or row.get("vol") or "0"
                    source_well = row.get("Source_Well") or row.get("source_well") or row.get("SourceWell") or ""
                    if well:
                        try:
                            vol = float(str(volume).replace(",", ".").replace("µL", "").replace("ul", "").strip())
                        except Exception:
                            vol = 0
                        operations.append({
                            "well": well.upper().strip(),
                            "liquid": liquid.strip(),
                            "volume": vol,
                            "source_well": source_well.upper().strip() if source_well else ""
                        })

            elif filename.endswith((".xlsx", ".xls")):
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
                    ws = wb.active
                    if ws is None:
                        return {"error": "Excel file has no active worksheet"}
                    headers = [str(cell.value or "").lower() for cell in ws[1]]
                    well_col = liquid_col = volume_col = source_col = None
                    for i, h in enumerate(headers):
                        if h in ["well", "destination", "dest", "target"]:
                            well_col = i
                        elif h in ["liquid", "source", "reagent", "sample"]:
                            liquid_col = i
                        elif h in ["volume", "vol", "volume_ul", "amount"]:
                            volume_col = i
                        elif h in ["source_well", "sourcewell", "from"]:
                            source_col = i
                    if well_col is None or volume_col is None:
                        return {"error": "Could not find 'Well' and 'Volume' columns in Excel file"}
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        well = str(row[well_col] or "") if len(row) > well_col else ""
                        liquid = str(row[liquid_col] or "") if liquid_col is not None and len(row) > liquid_col else ""
                        volume = row[volume_col] if len(row) > volume_col else 0
                        source_well = str(row[source_col] or "") if source_col is not None and len(row) > source_col else ""
                        if well:
                            try:
                                vol = float(str(volume).replace(",", ".").replace("µL", "").replace("ul", "").strip())
                            except Exception:
                                vol = 0
                            operations.append({
                                "well": well.upper().strip(),
                                "liquid": liquid.strip(),
                                "volume": vol,
                                "source_well": source_well.upper().strip() if source_well else ""
                            })
                except ImportError:
                    return {"error": "openpyxl not installed. Use CSV format or install: pip install openpyxl"}
            else:
                return {"error": "Unsupported file format. Use .csv or .xlsx"}

            if not operations:
                return {"error": "No valid operations found in file"}

            total_wells = len(operations)
            plates: list = []
            recipes: list = []
            timestamp = datetime.now().strftime("%Y%m%d")

            if total_wells <= plate_type or overflow_strategy == "multi_plate":
                num_plates = (total_wells + plate_type - 1) // plate_type
                for plate_idx in range(num_plates):
                    plate_id = f"{plate_prefix}-{timestamp}-{str(plate_idx + 1).zfill(3)}"
                    start = plate_idx * plate_type
                    end = min((plate_idx + 1) * plate_type, total_wells)
                    plate_ops = operations[start:end]
                    for op in plate_ops:
                        op["plate_id"] = plate_id
                    plates.append({"plate_id": plate_id, "plate_index": plate_idx, "well_count": len(plate_ops), "operations": plate_ops})
                recipe = generate_pipetting_recipe(plates, hal_config, plate_prefix, timestamp)
                recipes.append(recipe)
            else:
                num_recipes = (total_wells + plate_type - 1) // plate_type
                for recipe_idx in range(num_recipes):
                    plate_id = f"{plate_prefix}-{timestamp}-{str(recipe_idx + 1).zfill(3)}"
                    start = recipe_idx * plate_type
                    end = min((recipe_idx + 1) * plate_type, total_wells)
                    plate_ops = operations[start:end]
                    for i, op in enumerate(plate_ops):
                        op["plate_id"] = plate_id
                        op["original_well"] = op["well"]
                    plates.append({"plate_id": plate_id, "plate_index": recipe_idx, "well_count": len(plate_ops), "operations": plate_ops})
                    recipe = generate_pipetting_recipe([plates[-1]], hal_config, plate_prefix, timestamp)
                    recipe["name"] = f"Excel_Pipetting_{plate_id}"
                    recipes.append(recipe)

            async with plate_tracking_lock:
                for plate in plates:
                    plate_tracking[plate["plate_id"]] = {
                        "created": datetime.now().isoformat(),
                        "source_file": source_file_name,
                        "well_count": plate["well_count"],
                        "status": "pending",
                        "analysis_results": []
                    }
                await save_plate_tracking(plate_tracking)

            return {
                "total_wells": total_wells,
                "plate_capacity": plate_type,
                "num_plates": len(plates),
                "num_recipes": len(recipes),
                "overflow_strategy": overflow_strategy,
                "plates": [{"plate_id": p["plate_id"], "well_count": p["well_count"]} for p in plates],
                "operations": operations[:10],
                "recipes": recipes
            }

        except Exception as e:
            state.add_log("error", f"Excel parsing error: {str(e)}", "batch")
            return {"error": str(e)}

    @router.post("/parse-chemical-excel")
    async def parse_chemical_excel(request: Request):
        """
        Parse an Excel file and return detected columns for user configuration.

        Step 1: Parse → return columns and data
        Step 2: User configures groups/delays in UI
        Step 3: Call /api/batch/generate-recipe-with-config
        """
        from src.excel_recipe_parser import (
            parse_excel_columns,
            generate_default_config,
            generate_recipe_with_config,
            save_recipe_to_library
        )
        try:
            content_type = request.headers.get("content-type", "")

            if "application/json" in content_type:
                data = await request.json()
                library_file = data.get("library_file")
                config = data.get("config")
                save_to_library = data.get("save_to_library", False)
                if not library_file:
                    return {"error": "No library file specified"}
                file_path = LIBRARY_DIR / "Recipes" / "excel" / library_file
                if not file_path.exists():
                    return {"error": f"File not found: {library_file}"}
                with open(file_path, "rb") as f:
                    content = f.read()
                filename = library_file
            else:
                form = await request.form()
                file = form.get("file")
                config = None
                save_to_library = form.get("save_to_library", "false").lower() == "true"
                if not file:
                    return {"error": "No file uploaded"}
                content = await file.read()
                filename = file.filename or "unknown.xlsx"

            result = parse_excel_columns(content, filename)
            if "error" in result:
                return result

            for sheet in result.get("sheets", []):
                if "error" not in sheet:
                    sheet["default_config"] = generate_default_config(sheet)

            if config:
                recipes = []
                for sheet in result.get("sheets", []):
                    if "error" not in sheet:
                        recipe = generate_recipe_with_config(sheet, config)
                        if "error" not in recipe:
                            recipes.append(recipe)
                            if save_to_library:
                                save_recipe_to_library(recipe, LIBRARY_DIR)
                result["recipes"] = recipes

            state.add_log("info", f"Parsed Excel: {filename}, {len(result.get('sheets', []))} sheets", "batch")
            return result

        except Exception as e:
            state.add_log("error", f"Excel parsing error: {str(e)}", "batch")
            return {"error": str(e)}

    @router.post("/generate-recipe-with-config")
    async def generate_recipe_with_config_endpoint(request: Request):
        """Generate a recipe from parsed sheet data using user-defined group/delay config."""
        from src.excel_recipe_parser import (
            generate_recipe_with_config,
            save_recipe_to_library,
            generate_hal_config_from_recipe,
            save_hal_config_to_library
        )
        try:
            data = await request.json()
            sheet_data = data.get("sheet_data")
            config = data.get("config")
            recipe_name = data.get("recipe_name")
            save_to_library = data.get("save_to_library", False)
            generate_hal = data.get("generate_hal", False)

            if not sheet_data:
                return {"error": "No sheet_data provided"}
            if not config:
                return {"error": "No config provided"}

            recipe = generate_recipe_with_config(sheet_data, config, recipe_name)
            if "error" in recipe:
                return recipe

            result: dict = {"recipe": recipe}
            if save_to_library:
                result["saved_recipe"] = save_recipe_to_library(recipe, LIBRARY_DIR)
            if generate_hal:
                hal_config = generate_hal_config_from_recipe(recipe)
                result["hal_config"] = hal_config
                if save_to_library:
                    result["saved_hal"] = save_hal_config_to_library(hal_config, LIBRARY_DIR)

            state.add_log("info", f"Generated recipe: {recipe.get('ProtocolName')}", "batch")
            return result

        except Exception as e:
            state.add_log("error", f"Recipe generation error: {str(e)}", "batch")
            return {"error": str(e)}

    @router.get("/library-excel-files")
    async def list_library_excel_files():
        """List Excel files in Library/Recipes/excel for chemical recipe import."""
        files = []
        excel_dir = LIBRARY_DIR / "Recipes" / "excel"
        if excel_dir.exists():
            for f in sorted(excel_dir.glob("*.xlsx")) + sorted(excel_dir.glob("*.xls")):
                files.append({
                    "filename": f.name,
                    "path": str(f.relative_to(LIBRARY_DIR)),
                    "size": f.stat().st_size,
                    "modified": datetime.fromtimestamp(f.stat().st_mtime).isoformat(),
                    "type": "chemical"
                })
        return {"files": files, "directory": str(excel_dir)}

    @router.post("/generate-workflow")
    async def generate_workflow_from_recipes_api(request: Request):
        """Generate a workflow JSON from recipes with timing and extension points."""
        from src.excel_recipe_parser import generate_workflow_from_recipes, save_workflow_to_library
        try:
            data = await request.json()
            recipes = data.get("recipes", [])
            workflow_name = data.get("workflow_name")
            save_to_library = data.get("save_to_library", False)

            if not recipes:
                return {"error": "No recipes provided"}

            workflow = generate_workflow_from_recipes(recipes, workflow_name)
            saved_file = None
            if save_to_library:
                saved_file = save_workflow_to_library(workflow, LIBRARY_DIR)
                workflow["saved_file"] = saved_file

            state.add_log("info", f"Generated workflow '{workflow['name']}' with {len(workflow['steps'])} steps", "batch")
            return {"status": "success", "workflow": workflow, "saved_file": saved_file}

        except Exception as e:
            state.add_log("error", f"Workflow generation error: {str(e)}", "batch")
            return {"error": str(e)}

    @router.post("/generate-hal")
    async def generate_hal_config_api(request: Request):
        """Generate a generic HAL config from a recipe or explicit num_sources."""
        from src.excel_recipe_parser import generate_hal_config_from_recipe, generate_generic_hal, save_hal_config_to_library
        try:
            data = await request.json()
            recipe = data.get("recipe")
            hal_name = data.get("hal_name", "Generic_HAL")
            num_sources = data.get("num_sources")
            target_plate = data.get("target_plate", "corning_96_wellplate_360ul_flat")
            reservoir_type = data.get("reservoir_type", "nest_12_reservoir_15ml")
            save_to_library = data.get("save_to_library", False)

            if num_sources:
                hal_config = generate_generic_hal(num_sources=num_sources, hal_name=hal_name, target_plate=target_plate, reservoir_type=reservoir_type)
            elif recipe:
                hal_config = generate_hal_config_from_recipe(recipe, hal_name)
            else:
                hal_config = generate_generic_hal(num_sources=6, hal_name=hal_name, target_plate=target_plate, reservoir_type=reservoir_type)

            saved_file = None
            if save_to_library:
                saved_file = save_hal_config_to_library(hal_config, LIBRARY_DIR)
                hal_config["saved_file"] = saved_file

            state.add_log("info", f"Generated generic HAL '{hal_config.get('ConfigName')}' with {hal_config['Metadata']['total_sources']} sources", "batch")
            return {"status": "success", "hal_config": hal_config, "saved_file": saved_file}

        except Exception as e:
            state.add_log("error", f"HAL generation error: {str(e)}", "batch")
            return {"error": str(e)}

    return router
