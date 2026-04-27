#!/usr/bin/env python3
"""
Plug & Play Console
===================

A CLI that generates menus AUTOMATICALLY from SiLA2 server metadata.
NO HARDCODED INSTRUMENT MENUS - everything comes from the servers.

Features:
- Auto-discovery of all SiLA2 servers
- Dynamic menu generation from server features/commands
- Parameter prompts based on command definitions
- Workflow execution with any discovered instruments

To add a new instrument:
1. Create and start your SiLA2 server
2. The console automatically shows it with all its commands
3. NO CODE CHANGES NEEDED HERE

Usage:
    python pnp_console.py              # Interactive mode
    python pnp_console.py --discover   # Just show discovered servers
"""

import asyncio
import json
import os
import sys
from pathlib import Path
from typing import Any, Dict, List, Optional

# Base directory is where this script lives
BASE_DIR = Path(__file__).parent.absolute()
sys.path.insert(0, str(BASE_DIR))
sys.path.insert(0, str(BASE_DIR / "src"))  # For grpc imports

from src.discovery import PnPDiscovery, PnPServer, PnPCommand, PnPFeature
from src.client import PnPClient, PnPRegistry, CommandResult
from src.lab_core import LabCore, InterventionRequest, InterventionAction


#                              COLORS

class C:
    """ANSI color codes."""
    RESET = '\033[0m'
    BOLD = '\033[1m'
    DIM = '\033[2m'
    RED = '\033[91m'
    GREEN = '\033[92m'
    YELLOW = '\033[93m'
    BLUE = '\033[94m'
    MAGENTA = '\033[95m'
    CYAN = '\033[96m'
    WHITE = '\033[97m'


def clear():
    os.system('cls' if os.name == 'nt' else 'clear')


def ok(msg): print(f"{C.GREEN}[OK]{C.RESET} {msg}")
def err(msg): print(f"{C.RED}[ERR]{C.RESET} {msg}")
def info(msg): print(f"{C.CYAN}[i]{C.RESET} {msg}")
def warn(msg): print(f"{C.YELLOW}[!]{C.RESET} {msg}")


def print_header(title: str):
    print(f"\n{C.CYAN}{C.BOLD}{'='*60}{C.RESET}")
    print(f"{C.CYAN}{C.BOLD}  {title}{C.RESET}")
    print(f"{C.CYAN}{C.BOLD}{'='*60}{C.RESET}\n")


#                              MAIN APPLICATION

class PnPConsole:
    """
    Plug & Play Interactive Console.
    
    Generates all menus dynamically from server metadata.
    Uses LabCore as shared logic layer (same as WebApp).
    """
    
    def __init__(self):
        self.registry = PnPRegistry(BASE_DIR)
        self.lab_core = LabCore(BASE_DIR)  # Shared logic with WebApp
        self.running = True
        
        # Set up human intervention callback for workflow errors
        self.lab_core.set_intervention_callback(self._handle_intervention)
    
    async def _handle_intervention(self, request: InterventionRequest) -> InterventionAction:
        """
        Handle human intervention request from workflow executor.
        
        Shows interactive prompt to operator when a step fails.
        """
        print()
        print(f"\n{C.YELLOW}{'='*60}{C.RESET}")
        print(f"{C.YELLOW}  ⚠️  HUMAN INTERVENTION REQUIRED{C.RESET}")
        print(f"{C.YELLOW}{'='*60}{C.RESET}")
        print(f"\n  Workflow:   {request.workflow_name}")
        print(f"  Step:       {request.step_number}")
        print(f"  Instrument: {request.instrument}")
        print(f"  Action:     {request.action}")
        print(f"\n  {C.RED}Error:{C.RESET} {request.error.message}")
        print(f"  Category:   {request.error.category.value}")
        print(f"  Attempts:   {request.attempt}/{request.max_attempts}")
        
        if request.error.details:
            print(f"\n  Details: {request.error.details}")
        
        print(f"\n{C.YELLOW}{'='*60}{C.RESET}")
        print("\n  Options:")
        print(f"    {C.GREEN}[R]{C.RESET} Retry  - Fix the issue and retry this step")
        print(f"    {C.YELLOW}[S]{C.RESET} Skip   - Skip this step and continue workflow")
        print(f"    {C.RED}[A]{C.RESET} Abort  - Stop the entire workflow")
        print()
        
        while True:
            choice = input(f"  Your choice [{C.GREEN}R{C.RESET}/{C.YELLOW}S{C.RESET}/{C.RED}A{C.RESET}]: ").strip().upper()
            
            if choice == 'R':
                print(f"\n{C.GREEN}  → Retrying step...{C.RESET}\n")
                return InterventionAction.RETRY
            elif choice == 'S':
                print(f"\n{C.YELLOW}  → Skipping step...{C.RESET}\n")
                return InterventionAction.SKIP
            elif choice == 'A':
                print(f"\n{C.RED}  → Aborting workflow...{C.RESET}\n")
                return InterventionAction.ABORT
            else:
                print(f"  {C.RED}Invalid choice. Please enter R, S, or A.{C.RESET}")
    
    async def run(self):
        """Main application loop."""
        clear()
        print_header("PLUG & PLAY LAB CONSOLE")
        
        # Initial discovery
        info("Discovering SiLA2 servers...")
        count = await self.registry.discover()
        
        if count == 0:
            warn("No servers found!")
            print(f"\n{C.DIM}To add a server:{C.RESET}")
            print("  1. Create folder: SiLA2/YourInstrumentSiLA2Server/")
            print("  2. Add features: features/*.sila.xml")
            print("  3. Start your server")
            print()
        else:
            ok(f"Found {count} servers")
            
            # Auto-connect to online servers
            info("Connecting to online servers...")
            results = await self.registry.connect_all(timeout=3.0)
            online = sum(1 for v in results.values() if v)
            ok(f"Connected to {online}/{count} servers")
        
        # Main loop
        while self.running:
            try:
                await self.main_menu()
            except KeyboardInterrupt:
                print(f"\n\n{C.GREEN}Goodbye!{C.RESET}\n")
                break
            except Exception as e:
                import traceback
                err(f"Error: {e}")
                traceback.print_exc()
                await asyncio.sleep(1)
        
        # Cleanup
        await self.registry.disconnect_all()
    
    async def main_menu(self):
        """Main menu - dynamically lists all discovered servers."""
        clear()
        print(f"""
{C.CYAN}{C.BOLD}
  ____  _             ___     ____  _             
 |  _ \\| |_   _  __ _( _ )   |  _ \\| | __ _ _   _ 
 | |_) | | | | |/ _` / _ \\/\\ | |_) | |/ _` | | | |
 |  __/| | |_| | (_| | (_>  < |  __/| | (_| | |_| |
 |_|   |_|\\__,_|\\__, |\\___/\\/ |_|   |_|\\__,_|\\__, |
                |___/                        |___/ 
{C.RESET}
{C.DIM}        Plug & Play Lab Console v1.0{C.RESET}
{C.DIM}        Servers are the source of truth{C.RESET}
""")
        
        servers = self.registry.list_servers()
        
        print(f"{C.BOLD}Discovered Instruments:{C.RESET}\n")
        
        if not servers:
            print(f"  {C.DIM}No instruments found{C.RESET}")
            print(f"\n  {C.CYAN}r{C.RESET}. [R] Refresh")
            print(f"  {C.CYAN}0{C.RESET}. Exit")
        else:
            for i, server in enumerate(servers, 1):
                # Status indicator (ASCII-safe for Windows)
                if server.hardware_online:
                    status = f"{C.GREEN}[OK]{C.RESET}"
                elif server.server_online:
                    status = f"{C.YELLOW}[SW]{C.RESET}"  # Server only
                else:
                    status = f"{C.RED}[--]{C.RESET}"
                
                # Command count
                cmd_count = len(server.get_all_commands())
                
                # Type info
                type_info = server.server_type or server.vendor or ""
                if type_info:
                    type_info = f" [{type_info}]"
                
                print(f"  {C.CYAN}{i:2}{C.RESET}. {status} {server.name}{C.DIM}{type_info}{C.RESET}  ({cmd_count} commands)")
            
            print(f"\n{C.BOLD}Actions:{C.RESET}")
            print(f"  {C.CYAN}r{C.RESET}. [R] Refresh discovery")
            print(f"  {C.CYAN}w{C.RESET}. [W] Workflows")
            print(f"  {C.CYAN}b{C.RESET}. [B] Batch Execution")
            print(f"  {C.CYAN}s{C.RESET}. [S] System status")
            print(f"  {C.CYAN}0{C.RESET}. Exit")
        
        print()
        
        choice = input(f"{C.BOLD}Select: {C.RESET}").strip().lower()
        
        if choice == '0' or choice == 'q':
            self.running = False
        elif choice == 'r':
            await self._refresh()
        elif choice == 'w':
            await self.workflow_menu()
        elif choice == 'b':
            await self.batch_menu()
        elif choice == 's':
            await self.status_menu()
        elif choice.isdigit():
            idx = int(choice) - 1
            if 0 <= idx < len(servers):
                await self.server_menu(servers[idx])
    
    async def _refresh(self):
        """Refresh server discovery."""
        clear()
        print_header("REFRESHING...")
        
        await self.registry.disconnect_all()
        
        info("Scanning for servers...")
        count = await self.registry.discover()
        ok(f"Found {count} servers")
        
        info("Connecting...")
        results = await self.registry.connect_all(timeout=3.0)
        online = sum(1 for v in results.values() if v)
        ok(f"Connected to {online}/{count} servers")
        
        input(f"\n{C.DIM}Press Enter...{C.RESET}")
    
    #                         SERVER MENU (DYNAMIC)
    
    async def server_menu(self, server: PnPServer):
        """
        Dynamic menu for a server.
        
        Menu items are generated from the server's features/commands.
        NO HARDCODED ITEMS.
        """
        while True:
            clear()
            
            # Header with server info
            status = "[ONLINE]" if server.server_online else "[OFFLINE]"
            hw_status = server.hardware_status or "unknown"
            
            print_header(f"{server.name}")
            print(f"  Status: {status}")
            print(f"  Hardware: {hw_status}")
            print(f"  Address: {server.address}")
            if server.vendor:
                print(f"  Vendor: {server.vendor}")
            if server.serial_number:
                print(f"  Serial: {server.serial_number}")
            
            # Get all commands
            all_commands = server.get_all_commands()
            
            if not all_commands:
                print(f"\n  {C.DIM}No commands available{C.RESET}")
                print(f"\n  {C.CYAN}0{C.RESET}. <- Back")
                choice = input(f"\n{C.BOLD}Select: {C.RESET}").strip()
                if choice == '0':
                    break
                continue
            
            # Group commands by feature
            features_dict: Dict[str, List[tuple]] = {}
            for feature_id, cmd_id, cmd in all_commands:
                if feature_id not in features_dict:
                    features_dict[feature_id] = []
                features_dict[feature_id].append((cmd_id, cmd))
            
            # Display commands
            print(f"\n{C.BOLD}Commands:{C.RESET}\n")
            
            menu_items = []
            idx = 1
            
            for feature_id, commands in features_dict.items():
                # Find feature display name
                feature_name = feature_id
                for f in server.features:
                    if f.identifier == feature_id:
                        feature_name = f.display_name or feature_id
                        break
                
                print(f"  {C.DIM}-- {feature_name} --{C.RESET}")
                
                for cmd_id, cmd in commands:
                    param_hint = ""
                    if cmd.parameters:
                        param_names = [p.identifier for p in cmd.parameters[:2]]
                        param_hint = f" {C.DIM}({', '.join(param_names)}{'...' if len(cmd.parameters) > 2 else ''}){C.RESET}"
                    
                    streaming = " [stream]" if cmd.observable else ""
                    
                    print(f"  {C.CYAN}{idx:2}{C.RESET}. {cmd.display_name or cmd_id}{streaming}{param_hint}")
                    menu_items.append((feature_id, cmd_id, cmd))
                    idx += 1
            
            print(f"\n  {C.CYAN} 0{C.RESET}. <- Back")
            print()
            
            choice = input(f"{C.BOLD}Select: {C.RESET}").strip()
            
            if choice == '0' or choice.lower() == 'q':
                break
            
            if choice.isdigit():
                cmd_idx = int(choice) - 1
                if 0 <= cmd_idx < len(menu_items):
                    feature_id, cmd_id, cmd = menu_items[cmd_idx]
                    await self.execute_command(server, feature_id, cmd)
    
    #                    LIBRARY FILE SELECTION HELPERS
    
    # Map of (command_identifier, parameter_identifier) -> Library subfolder
    LIBRARY_FILE_PARAMS = {
        # Opentrons recipe execution - all variations
        ("ExecuteRecipe", "RecipeName"): "Recipes",
        ("ExecuteRecipe", "recipe_name"): "Recipes",
        ("ExecuteRecipe", "recipe_file"): "Recipes",
        ("ExecuteRecipe", "RecipeFile"): "Recipes",
        ("ExecuteRecipeByName", "RecipeName"): "Recipes",
        ("ExecuteRecipeByName", "recipe_name"): "Recipes",
        ("LoadRecipe", "recipe_file"): "Recipes",
        ("LoadRecipe", "RecipeFile"): "Recipes",
        ("RunProtocol", "json_content"): "Recipes",  # For JSON protocol content
        # Protocol/measurement execution (Tecan etc)
        ("RunMeasurement", "protocol_file"): "Analysis",
        ("RunMeasurement", "ProtocolFile"): "Analysis",
        ("RunMeasurement", "ProtocolName"): "Analysis",
        ("RunProtocol", "protocol_file"): "Analysis",
        ("RunProtocol", "ProtocolFile"): "Analysis",
        ("RunProtocol", "ProtocolName"): "Analysis",
        # Hardware configuration
        ("SwitchHardwareConfig", "config_file"): "HardwareConfig",
        ("SwitchHardwareConfig", "ConfigFile"): "HardwareConfig",
        ("SwitchHardwareConfig", "ConfigName"): "HardwareConfig",
        ("SwitchHardwareConfig", "config_name"): "HardwareConfig",
        ("LoadHardwareConfig", "config_file"): "HardwareConfig",
        ("LoadHardwareConfig", "ConfigFile"): "HardwareConfig",
        ("LoadHardwareConfig", "ConfigName"): "HardwareConfig",
        # Workflow files
        ("LoadWorkflow", "workflow_file"): "Workflows",
        ("ExecuteWorkflow", "workflow_file"): "Workflows",
        ("ExecuteWorkflow", "WorkflowName"): "Workflows",
    }
    
    def _get_library_folder_for_param(self, cmd_id: str, param_id: str) -> str:
        """
        Check if this parameter should offer file selection from Library.
        Returns Library subfolder name, or empty string if not applicable.
        """
        # Check exact match first
        key = (cmd_id, param_id)
        if key in self.LIBRARY_FILE_PARAMS:
            return self.LIBRARY_FILE_PARAMS[key]
        
        # Check by parameter name patterns (case insensitive)
        param_lower = param_id.lower()
        cmd_lower = cmd_id.lower()
        
        # Recipe patterns
        if "recipe" in param_lower or "recipe" in cmd_lower:
            if any(x in param_lower for x in ["name", "file", "path", "recipe"]):
                return "Recipes"
        
        # Protocol/Analysis patterns
        if "protocol" in param_lower or "measurement" in param_lower or "analysis" in param_lower:
            return "Analysis"
        if "protocol" in cmd_lower and any(x in param_lower for x in ["name", "file", "path"]):
            return "Analysis"
        
        # Hardware config patterns
        if "config" in param_lower and ("hardware" in param_lower or "hardware" in cmd_lower or "hal" in param_lower):
            return "HardwareConfig"
        if "switchhardware" in cmd_lower or "loadhardware" in cmd_lower:
            return "HardwareConfig"
        
        # Workflow patterns
        if "workflow" in param_lower or "workflow" in cmd_lower:
            if any(x in param_lower for x in ["name", "file", "path", "workflow"]):
                return "Workflows"
        
        return ""
    
    def _select_library_file(self, subfolder: str, show_full_path: bool = False) -> str:
        """
        Show available files from Library subfolder and let user select.
        Returns selected filename, or empty string if cancelled.
        """
        library_path = BASE_DIR / "Library" / subfolder
        
        if not library_path.exists():
            print(f"  {C.YELLOW}Warning: Library/{subfolder} not found{C.RESET}")
            return ""
        
        # Get available files
        files = []
        for f in sorted(library_path.iterdir()):
            if f.is_file() and not f.name.startswith('.'):
                files.append(f)
        
        if not files:
            print(f"  {C.YELLOW}No files in Library/{subfolder}{C.RESET}")
            return ""
        
        print(f"\n  {C.BOLD}Available files in Library/{subfolder}:{C.RESET}")
        for i, f in enumerate(files, 1):
            # Show file info
            size = f.stat().st_size
            size_str = f"{size/1024:.1f}KB" if size > 1024 else f"{size}B"
            print(f"    {C.CYAN}{i:2}{C.RESET}. {f.name} {C.DIM}({size_str}){C.RESET}")
        
        print(f"    {C.DIM} 0. (manual input){C.RESET}")
        
        choice = input(f"\n  {C.BOLD}Select file [1-{len(files)}]: {C.RESET}").strip()
        
        if choice == '0':
            return ""  # User wants manual input
        
        if choice.isdigit():
            idx = int(choice) - 1
            if 0 <= idx < len(files):
                selected = files[idx]
                if show_full_path:
                    return str(selected)
                return selected.name
        
        return ""  # Invalid choice, fall back to manual input
    
    async def _get_tip_racks_from_server(self, server: PnPServer) -> list:
        """
        Query server for available tip racks from HAL config.
        Returns list of dicts with 'id', 'type', 'used', 'available'.
        """
        try:
            # Use direct gRPC call for reliability
            import grpc
            from src.pnp_stubs import OpentronsService_pb2 as pb2
            from src.pnp_stubs import OpentronsService_pb2_grpc as pb2_grpc
            from google.protobuf import empty_pb2
            
            async with grpc.aio.insecure_channel(server.address) as channel:
                stub = pb2_grpc.OpentronsServiceStub(channel)
                response = await stub.ListTipRacks(empty_pb2.Empty())
                
                if response.tip_racks_json:
                    import json
                    parsed = json.loads(response.tip_racks_json)
                    return parsed.get("racks", [])
        except Exception as e:
            # Silently fail - will fall back to manual input
            pass
        return []
    
    async def execute_command(self, server: PnPServer, feature_id: str, cmd: PnPCommand):
        """
        Execute a command with dynamic parameter input.
        
        Parameters are prompted based on the command definition.
        """
        clear()
        print_header(f"Execute: {cmd.display_name or cmd.identifier}")
        
        # Collect parameters
        params = {}
        
        if cmd.parameters:
            print(f"{C.BOLD}Parameters:{C.RESET}\n")
            
            for param in cmd.parameters:
                # Show parameter info
                required = "*" if param.required else ""
                type_hint = f" [{param.data_type}]" if param.data_type else ""
                default_hint = f" (default: {param.default_value})" if param.default_value else ""
                
                # Check if this parameter should offer Library file selection
                library_folder = self._get_library_folder_for_param(cmd.identifier, param.identifier)
                
                value = ""
                if library_folder:
                    # Offer file selection from Library
                    value = self._select_library_file(library_folder)
                
                # Special handling for RefillTipRack.RackType - query server for HAL tip racks
                if not value and cmd.identifier == "RefillTipRack" and "rack" in param.identifier.lower():
                    tip_racks = await self._get_tip_racks_from_server(server)
                    if tip_racks:
                        print(f"\n  {C.BOLD}Available Tip Racks (from HAL config):{C.RESET}")
                        for i, rack in enumerate(tip_racks, 1):
                            rack_id = rack.get("id", rack.get("type", "unknown"))
                            rack_type = rack.get("type", "")
                            
                            # Special display for "all" option
                            if rack_type == "all":
                                print(f"    {C.YELLOW}{i:2}{C.RESET}. {C.BOLD}REFILL ALL{C.RESET} {C.DIM}(reset all tip racks){C.RESET}")
                            else:
                                used = rack.get("used", 0)
                                avail = rack.get("available", 96)
                                # Color code based on availability
                                if avail == 0:
                                    status_color = C.RED
                                    status = "EMPTY"
                                elif avail < 20:
                                    status_color = C.YELLOW
                                    status = f"{avail} left"
                                else:
                                    status_color = C.GREEN
                                    status = f"{avail} left"
                                print(f"    {C.CYAN}{i:2}{C.RESET}. {rack_id} {C.DIM}({rack_type}){C.RESET}")
                                print(f"        {status_color}[{status}]{C.RESET} {C.DIM}Used: {used}/96{C.RESET}")
                        print(f"    {C.DIM} 0. (manual input){C.RESET}")
                        
                        choice = input(f"\n  Select [1-{len(tip_racks)}]: ").strip()
                        if choice.isdigit() and 1 <= int(choice) <= len(tip_racks):
                            # Use the LoadName (type) for refill
                            value = tip_racks[int(choice) - 1].get("type", "")
                
                if not value and param.constraints:
                    # Offer selection from constraints
                    print(f"\n  {C.BOLD}{param.display_name or param.identifier}:{C.RESET}")
                    for i, opt in enumerate(param.constraints, 1):
                        print(f"    {C.CYAN}{i:2}{C.RESET}. {opt}")
                    print(f"    {C.DIM} 0. (manual input){C.RESET}")
                    
                    choice = input(f"\n  Select [1-{len(param.constraints)}]: ").strip()
                    if choice.isdigit() and 1 <= int(choice) <= len(param.constraints):
                        value = param.constraints[int(choice) - 1]
                    elif choice == '0':
                        value = ""  # Fall through to manual input
                    else:
                        value = ""
                
                if not value:
                    # Manual input
                    prompt = f"  {param.display_name or param.identifier}{required}{type_hint}{default_hint}: "
                    value = input(prompt).strip()
                
                # Use default if empty
                if not value and param.default_value:
                    value = param.default_value
                
                # Skip empty optional parameters
                if not value and not param.required:
                    continue
                
                # Type conversion
                if param.data_type == "Integer":
                    try:
                        value = int(value)
                    except ValueError:
                        pass
                elif param.data_type == "Real":
                    try:
                        value = float(value)
                    except ValueError:
                        pass
                elif param.data_type == "Boolean":
                    value = value.lower() in ("true", "yes", "1", "on")
                
                params[param.identifier] = value
        
        # Confirm execution
        print(f"\n{C.BOLD}Executing:{C.RESET}")
        print(f"  Server: {server.name}")
        print(f"  Command: {cmd.identifier}")
        if params:
            print(f"  Parameters: {params}")
        
        # Special handling for RefillTipRack - require operator physical confirmation
        if cmd.identifier == "RefillTipRack":
            rack_type = params.get("RackType", params.get("rack_type", "unknown"))
            print()
            print(f"{C.RED}{'═' * 60}{C.RESET}")
            print(f"{C.BOLD}{C.YELLOW}  ⚠ OPERATOR ACTION REQUIRED ⚠{C.RESET}")
            print(f"{C.RED}{'═' * 60}{C.RESET}")
            print()
            print(f"  {C.BOLD}Please refill the tip rack:{C.RESET}")
            print(f"  {C.CYAN}{rack_type}{C.RESET}")
            print()
            print(f"  1. Open the Opentrons door")
            print(f"  2. Replace or refill the tip rack")
            print(f"  3. Close the door")
            print()
            print(f"{C.RED}{'─' * 60}{C.RESET}")
            confirm = input(f"\n{C.BOLD}Press ENTER when refill is complete (or 'cancel'): {C.RESET}").strip().lower()
            if confirm == 'cancel':
                warn("Refill cancelled")
                return
            ok("Refill confirmed by operator")
        else:
            confirm = input(f"\n{C.BOLD}Execute? [Y/n]: {C.RESET}").strip().lower()
            if confirm and confirm != 'y':
                return
        
        # Execute
        print(f"\n{C.CYAN}Executing...{C.RESET}")
        
        # Use LabCore for execution (shared logic with WebApp)
        # Map server.name to instrument_id for LabCore
        instrument_id = server.name.lower().replace(" ", "_")
        
        # Ensure LabCore has discovered instruments
        await self.lab_core.discover()
        
        result = await self.lab_core.execute_command(instrument_id, cmd.identifier, params)
        
        # Show result
        print()
        if result.success:
            ok("Command completed successfully!")
            if result.data:
                print(f"\n{C.BOLD}Result:{C.RESET}")
                if isinstance(result.data, dict):
                    for key, value in result.data.items():
                        if key not in ('success',):
                            print(f"  {key}: {value}")
                else:
                    print(f"  {result.data}")
        else:
            err(f"Command failed: {result.error}")
        
        input(f"\n{C.DIM}Press Enter...{C.RESET}")
    
    #                         WORKFLOW MENU
    
    async def workflow_menu(self):
        """Workflow management menu - similar UX to instrument selection."""
        while True:
            clear()
            print_header("WORKFLOW MANAGEMENT")
            
            # List workflow files
            workflows_dir = BASE_DIR / "Library" / "Workflows"
            workflow_files = []
            
            if workflows_dir.exists():
                workflow_files = sorted(workflows_dir.glob("*.workflow.json"))
            
            print(f"{C.BOLD}Available Workflows:{C.RESET}\n")
            
            if not workflow_files:
                print(f"  {C.DIM}No workflows found in Library/Workflows/{C.RESET}")
            else:
                for i, wf_path in enumerate(workflow_files, 1):
                    # Load workflow to show summary
                    try:
                        with open(wf_path) as f:
                            wf = json.load(f)
                        steps = wf.get("Steps", [])
                        desc = wf.get("Description", "")[:40]
                        
                        # Detect instruments used
                        instruments = set(s.get("Instrument", "") for s in steps)
                        inst_str = ", ".join(instruments) if instruments else "N/A"
                        
                        print(f"  {C.CYAN}{i:2}{C.RESET}. {wf_path.stem}")
                        print(f"      {C.DIM}{len(steps)} steps | {inst_str}{C.RESET}")
                        if desc:
                            print(f"      {C.DIM}{desc}...{C.RESET}")
                    except Exception:
                        print(f"  {C.CYAN}{i:2}{C.RESET}. {wf_path.stem} {C.RED}(error loading){C.RESET}")
            
            print(f"\n{C.BOLD}Actions:{C.RESET}")
            print(f"  {C.CYAN}c{C.RESET}. Create new workflow")
            print(f"  {C.CYAN}0{C.RESET}. <- Back")
            print()
            
            choice = input(f"{C.BOLD}Select workflow [1-{len(workflow_files)}]: {C.RESET}").strip().lower()
            
            if choice == '0' or choice == 'q':
                break
            elif choice == 'c':
                await self.create_workflow()
            elif choice.isdigit():
                idx = int(choice) - 1
                if 0 <= idx < len(workflow_files):
                    await self.workflow_detail_menu(workflow_files[idx])
    
    async def workflow_detail_menu(self, workflow_path: Path):
        """Show workflow details and allow execution - similar to instrument command menu."""
        try:
            with open(workflow_path) as f:
                workflow = json.load(f)
        except Exception as e:
            err(f"Failed to load workflow: {e}")
            input(f"\n{C.DIM}Press Enter...{C.RESET}")
            return
        
        while True:
            clear()
            print_header(f"Workflow: {workflow.get('WorkflowName', workflow_path.stem)}")
            
            if workflow.get("Description"):
                print(f"{C.DIM}{workflow['Description']}{C.RESET}\n")
            
            steps = workflow.get("Steps", [])
            
            # Show steps overview
            print(f"{C.BOLD}Steps ({len(steps)}):{C.RESET}\n")
            
            has_opentrons = False
            has_manual = False
            
            for step in steps:
                num = step.get("StepNumber", "?")
                inst = step.get("Instrument", "Unknown")
                action = step.get("Action", "Unknown")
                desc = step.get("Description", "")[:50]
                
                # Detect special instruments
                if "opentrons" in inst.lower():
                    has_opentrons = True
                if "manual" in inst.lower():
                    has_manual = True
                
                # Format step
                print(f"  {C.CYAN}{num:2}{C.RESET}. [{inst}] {action}")
                if desc:
                    print(f"      {C.DIM}{desc}{C.RESET}")
                
                # Show key parameters
                params = step.get("Parameters", {})
                if params:
                    param_str = ", ".join(f"{k}={v}" for k, v in list(params.items())[:3])
                    print(f"      {C.DIM}Params: {param_str}{C.RESET}")
            
            # Warnings/notes
            print()
            if has_opentrons:
                print(f"  {C.YELLOW}⚠ Contains Opentrons steps - HAL config will be requested{C.RESET}")
            if has_manual:
                print(f"  {C.YELLOW}⚠ Contains Manual Station steps - operator confirmation required{C.RESET}")
            
            print(f"\n{C.BOLD}Actions:{C.RESET}")
            print(f"  {C.GREEN}r{C.RESET}. [R]un workflow")
            print(f"  {C.CYAN}v{C.RESET}. [V]iew full JSON")
            print(f"  {C.CYAN}0{C.RESET}. <- Back")
            print()
            
            choice = input(f"{C.BOLD}Select: {C.RESET}").strip().lower()
            
            if choice == '0' or choice == 'q':
                break
            elif choice == 'r':
                await self.run_workflow_with_hal(workflow_path, workflow)
            elif choice == 'v':
                clear()
                print_header("Workflow JSON")
                print(json.dumps(workflow, indent=2))
                input(f"\n{C.DIM}Press Enter...{C.RESET}")
    
    async def run_workflow_with_hal(self, workflow_path: Path, workflow: dict):
        """Run workflow with HAL selection for Opentrons steps and manual confirmation."""
        clear()
        print_header(f"Execute: {workflow.get('WorkflowName', workflow_path.stem)}")
        
        steps = workflow.get("Steps", [])
        
        # Check if workflow has Opentrons steps that need HAL
        opentrons_steps = [s for s in steps if "opentrons" in s.get("Instrument", "").lower()]
        hal_config = None
        
        if opentrons_steps:
            print(f"{C.BOLD}This workflow contains Opentrons steps.{C.RESET}")
            print(f"Please select a Hardware Configuration (HAL):\n")
            
            # List available HAL configs
            hal_dir = BASE_DIR / "Library" / "HardwareConfig"
            hal_files = []
            if hal_dir.exists():
                hal_files = sorted(hal_dir.glob("*.json"))
            
            if hal_files:
                for i, hf in enumerate(hal_files, 1):
                    print(f"  {C.CYAN}{i}{C.RESET}. {hf.stem}")
                print(f"  {C.DIM}0. Skip HAL (use recipe defaults){C.RESET}")
                
                choice = input(f"\n{C.BOLD}Select HAL config [1-{len(hal_files)}]: {C.RESET}").strip()
                
                if choice.isdigit() and 1 <= int(choice) <= len(hal_files):
                    hal_config = hal_files[int(choice) - 1].name
                    ok(f"Using HAL: {hal_config}")
                else:
                    warn("No HAL selected - using recipe defaults")
            else:
                warn("No HAL configs found in Library/HardwareConfig/")
            
            print()
        
        # Confirmation
        print(f"{C.BOLD}Ready to execute {len(steps)} steps.{C.RESET}\n")
        confirm = input(f"Execute workflow? [Y/n]: ").strip().lower()
        if confirm and confirm != 'y':
            return
        
        # Ensure LabCore is ready for execution
        await self.lab_core.discover()
        
        # Execute steps
        print(f"\n{C.BOLD}{'═' * 60}{C.RESET}")
        print(f"{C.BOLD}  EXECUTING WORKFLOW{C.RESET}")
        print(f"{C.BOLD}{'═' * 60}{C.RESET}\n")
        
        for step in steps:
            step_num = step.get("StepNumber", "?")
            instrument = step.get("Instrument", "")
            action = step.get("Action", "")
            params = step.get("Parameters", {}).copy()
            desc = step.get("Description", "")
            
            print(f"{C.BOLD}[Step {step_num}]{C.RESET} {instrument} -> {action}")
            if desc:
                print(f"  {C.DIM}{desc}{C.RESET}")
            
            # Handle Manual Station specially
            if "manual" in instrument.lower():
                success = await self._handle_manual_step(step)
                if not success:
                    err("  Manual operation cancelled or failed")
                    cont = input(f"\n{C.BOLD}Continue workflow? [y/N]: {C.RESET}").strip().lower()
                    if cont != 'y':
                        break
                continue
            
            # Inject HAL config for Opentrons steps
            if "opentrons" in instrument.lower() and hal_config:
                params["HALConfig"] = hal_config
                params["UseHAL"] = True
            
            # Map instrument name to LabCore instrument_id
            instrument_id = instrument.lower().replace(" ", "_")
            
            # Execute via LabCore (same as single command) - this properly waits for completion
            print(f"  {C.CYAN}Executing...{C.RESET}", end=" ", flush=True)
            result = await self.lab_core.execute_command(instrument_id, action, params)
            
            # Check for errors - both in result.success AND in result.data.status
            step_failed = False
            error_message = ""
            
            if not result.success:
                step_failed = True
                error_message = result.error or "Unknown error"
            elif result.data and isinstance(result.data, dict):
                # Check status field in response data (Opentrons returns status: error)
                data_status = result.data.get("status", "").lower()
                if data_status == "error":
                    step_failed = True
                    error_message = result.data.get("error_message") or result.data.get("message") or "Operation failed on device"
            
            if step_failed:
                err(f"FAILED")
                print(f"    {C.RED}Error: {error_message}{C.RESET}")
                
                # For Opentrons steps, offer recovery options
                if "opentrons" in instrument.lower():
                    print()
                    print(f"{C.YELLOW}{'─' * 60}{C.RESET}")
                    print(f"  {C.BOLD}{C.YELLOW}⚠️  OPENTRONS ERROR - RECOVERY OPTIONS{C.RESET}")
                    print(f"{C.YELLOW}{'─' * 60}{C.RESET}")
                    print(f"  {C.CYAN}1{C.RESET}. Refill consumables and RETRY this step")
                    print(f"  {C.CYAN}2{C.RESET}. SKIP this step and continue workflow")
                    print(f"  {C.RED}3{C.RESET}. ABORT workflow")
                    print()
                    
                    choice = input(f"{C.BOLD}Choose [1-3]: {C.RESET}").strip()
                    
                    if choice == "1":
                        # Retry - operator refills consumables
                        print(f"\n  {C.CYAN}Waiting for operator to refill...{C.RESET}")
                        
                        # Try to extract rack type from error message
                        extracted_rack = None
                        if "tips" in error_message.lower() or "rack" in error_message.lower():
                            # Look for rack type pattern like "opentrons_flex_96_tiprack_1000ul"
                            import re
                            rack_match = re.search(r'(opentrons_flex_96_\w+|opentrons_flex_96_filtertiprack_\w+)', error_message)
                            if rack_match:
                                extracted_rack = rack_match.group(1)
                                print(f"  {C.DIM}Detected rack type: {extracted_rack}{C.RESET}")
                        
                        input(f"  {C.GREEN}Press ENTER when ready to retry{C.RESET}")
                        
                        # Reset tip tracker on server before retry
                        print(f"  {C.CYAN}Resetting tip tracker...{C.RESET}", end=" ", flush=True)
                        # Use specific rack if detected, otherwise reset all
                        rack_to_reset = extracted_rack or "all"
                        reset_result = await self.lab_core.execute_command(
                            instrument_id, 
                            "RefillTipRack", 
                            {"rack_type": rack_to_reset}
                        )
                        if reset_result.success:
                            ok("Done")
                        else:
                            warn(f"Reset failed (continuing anyway)")
                        
                        print(f"  {C.CYAN}Retrying...{C.RESET}", end=" ", flush=True)
                        result = await self.lab_core.execute_command(instrument_id, action, params)
                        
                        # Check retry result
                        if result.success and (not result.data or result.data.get("status", "").lower() != "error"):
                            ok(f"Completed on retry")
                            if result.data and isinstance(result.data, dict):
                                for k, v in list(result.data.items())[:3]:
                                    print(f"    {C.DIM}{k}: {v}{C.RESET}")
                        else:
                            err(f"Retry also failed")
                            retry_error = ""
                            if result.data and isinstance(result.data, dict):
                                retry_error = result.data.get("message") or result.error or ""
                            if retry_error:
                                print(f"    {C.RED}{retry_error}{C.RESET}")
                            break
                    elif choice == "2":
                        warn("Step skipped by operator")
                    else:
                        err("Workflow aborted by operator")
                        break
                else:
                    # Non-Opentrons error
                    cont = input(f"\n{C.BOLD}Continue workflow? [y/N]: {C.RESET}").strip().lower()
                    if cont != 'y':
                        break
            else:
                ok(f"Completed")
                if result.data and isinstance(result.data, dict):
                    for k, v in list(result.data.items())[:3]:
                        print(f"    {C.DIM}{k}: {v}{C.RESET}")
            
            print()
        
        print(f"{C.BOLD}{'═' * 60}{C.RESET}")
        ok("Workflow finished")
        input(f"\n{C.DIM}Press Enter...{C.RESET}")
    
    async def _handle_manual_step(self, step: dict) -> bool:
        """
        Handle a Manual Station step with terminal confirmation.
        
        Returns True if operator confirmed, False if cancelled.
        """
        params = step.get("Parameters", {})
        task_type = params.get("task_type", "manual_operation")
        description = params.get("description", step.get("Description", "Manual operation required"))
        priority = params.get("priority", "normal")
        timeout = params.get("timeout_seconds", 300)
        
        # Display manual operation request
        print()
        print(f"{C.RED}{'═' * 60}{C.RESET}")
        print(f"{C.BOLD}{C.RED}  🔔 OPERATOR ACTION REQUIRED 🔔{C.RESET}")
        print(f"{C.RED}{'═' * 60}{C.RESET}")
        print()
        print(f"  {C.BOLD}Type:{C.RESET}     {task_type}")
        print(f"  {C.BOLD}Priority:{C.RESET} {priority.upper()}")
        print()
        print(f"  {C.BOLD}OPERATION:{C.RESET}")
        print(f"  {C.YELLOW}{description}{C.RESET}")
        print()
        print(f"{C.RED}{'─' * 60}{C.RESET}")
        print(f"  {C.GREEN}Press ENTER when operation is complete{C.RESET}")
        print(f"  {C.RED}Type 'skip' to skip, 'cancel' to cancel workflow{C.RESET}")
        print(f"{C.RED}{'─' * 60}{C.RESET}")
        
        response = input(f"\n{C.BOLD}Confirm: {C.RESET}").strip().lower()
        
        if response == 'cancel':
            return False
        elif response == 'skip':
            warn("  Step skipped by operator")
            return True
        else:
            ok("  Operator confirmed")
            return True
    
    async def run_workflow(self, workflow_path: Path):
        """Run a workflow file - redirects to new detailed method."""
        try:
            with open(workflow_path) as f:
                workflow = json.load(f)
            await self.run_workflow_with_hal(workflow_path, workflow)
        except Exception as e:
            err(f"Failed to load workflow: {e}")
            input(f"\n{C.DIM}Press Enter...{C.RESET}")
    
    async def create_workflow(self):
        """Interactive workflow builder."""
        clear()
        print_header("CREATE WORKFLOW")
        
        servers = self.registry.get_online_servers()
        if not servers:
            warn("No online servers - cannot create workflow")
            input(f"\n{C.DIM}Press Enter...{C.RESET}")
            return
        
        workflow_name = input(f"{C.BOLD}Workflow name: {C.RESET}").strip()
        if not workflow_name:
            return
        
        steps = []
        step_num = 1
        
        print(f"\n{C.DIM}Add steps (empty line to finish){C.RESET}\n")
        
        while True:
            print(f"{C.BOLD}Step {step_num}:{C.RESET}")
            
            # Select instrument
            print("  Available instruments:")
            for i, s in enumerate(servers, 1):
                print(f"    {i}. {s.name}")
            
            inst_choice = input("  Instrument number (or empty to finish): ").strip()
            if not inst_choice:
                break
            
            try:
                inst_idx = int(inst_choice) - 1
                if inst_idx < 0 or inst_idx >= len(servers):
                    continue
            except ValueError:
                continue
            
            server = servers[inst_idx]
            
            # Select command
            commands = server.get_all_commands()
            print(f"  Commands for {server.name}:")
            for i, (_, cmd_id, _) in enumerate(commands, 1):
                print(f"    {i}. {cmd_id}")
            
            cmd_choice = input("  Command number: ").strip()
            try:
                cmd_idx = int(cmd_choice) - 1
                if cmd_idx < 0 or cmd_idx >= len(commands):
                    continue
            except ValueError:
                continue
            
            feature_id, cmd_id, cmd = commands[cmd_idx]
            
            # Get parameters (with Library file selection like execute_command)
            params = {}
            if cmd.parameters:
                print("  Parameters:")
                for param in cmd.parameters:
                    # Check if this parameter should offer Library file selection
                    library_folder = self._get_library_folder_for_param(cmd_id, param.identifier)
                    
                    value = ""
                    if library_folder:
                        # Offer file selection from Library
                        value = self._select_library_file(library_folder)
                    
                    if not value and param.constraints:
                        # Offer selection from constraints
                        print(f"\n    {C.BOLD}{param.display_name or param.identifier}:{C.RESET}")
                        for i, opt in enumerate(param.constraints, 1):
                            print(f"      {C.CYAN}{i:2}{C.RESET}. {opt}")
                        print(f"      {C.DIM} 0. (manual input){C.RESET}")
                        
                        choice = input(f"\n    Select [1-{len(param.constraints)}]: ").strip()
                        if choice.isdigit() and 1 <= int(choice) <= len(param.constraints):
                            value = param.constraints[int(choice) - 1]
                    
                    if not value:
                        # Manual input
                        value = input(f"    {param.identifier}: ").strip()
                    
                    if value:
                        params[param.identifier] = value
            
            steps.append({
                "StepNumber": step_num,
                "Instrument": server.name,
                "Action": cmd_id,
                "Parameters": params
            })
            
            step_num += 1
            print()
        
        if not steps:
            return
        
        # Save workflow
        workflow = {
            "WorkflowName": workflow_name,
            "Steps": steps
        }
        
        workflows_dir = BASE_DIR / "Library" / "Workflows"
        workflows_dir.mkdir(parents=True, exist_ok=True)
        
        filename = f"{workflow_name.replace(' ', '_')}.workflow.json"
        filepath = workflows_dir / filename
        
        with open(filepath, 'w') as f:
            json.dump(workflow, f, indent=2)
        
        ok(f"Saved: {filepath}")
        input(f"\n{C.DIM}Press Enter...{C.RESET}")
    
    #                         BATCH EXECUTION
    
    async def batch_menu(self):
        """Batch execution menu - run multiple workflows/recipes in sequence."""
        batch_queue = []  # List of (type, name, data) - type: 'workflow' | 'generated_recipe'
        
        while True:
            clear()
            print_header("BATCH EXECUTION")
            
            # Show current queue
            print(f"{C.BOLD}Queue ({len(batch_queue)} items):{C.RESET}\n")
            
            if not batch_queue:
                print(f"  {C.DIM}Queue is empty{C.RESET}")
            else:
                for i, (item_type, name, data) in enumerate(batch_queue, 1):
                    if item_type == 'workflow':
                        steps = len(data.get("Steps", []))
                        print(f"  {C.CYAN}{i:2}{C.RESET}. [WF] {name} ({steps} steps)")
                    elif item_type == 'generated_recipe':
                        ops = data.get("metadata", {}).get("total_operations", 0)
                        plate = data.get("plate_ids", ["?"])[0]
                        print(f"  {C.GREEN}{i:2}{C.RESET}. [RC] {name} (plate: {plate}, {ops} ops)")
            
            print(f"\n{C.BOLD}Actions:{C.RESET}")
            print(f"  {C.CYAN}a{C.RESET}. [A] Add workflow to queue")
            print(f"  {C.CYAN}r{C.RESET}. [R] Add recipe to queue")
            print(f"  {C.GREEN}e{C.RESET}. [E] Import from Excel/CSV (genera ricette)")
            print(f"  {C.CYAN}c{C.RESET}. [C] Clear queue")
            if batch_queue:
                print(f"  {C.GREEN}x{C.RESET}. [X] Execute batch")
            print(f"  {C.CYAN}0{C.RESET}. Back")
            
            print()
            choice = input(f"{C.BOLD}Select: {C.RESET}").strip().lower()
            
            if choice == '0' or choice == 'q':
                break
            elif choice == 'a':
                # Add workflow to queue
                selected = await self._select_workflow_for_batch()
                if selected:
                    path, wf_data = selected
                    batch_queue.append(('workflow', wf_data.get('WorkflowName', path.stem), wf_data))
                    ok(f"Added workflow: {wf_data.get('WorkflowName', path.stem)}")
                    input(f"\n{C.DIM}Press Enter...{C.RESET}")
            elif choice == 'r':
                # Add recipe to queue
                selected = await self._select_recipe_for_batch()
                if selected:
                    path, recipe_data = selected
                    batch_queue.append(('recipe', recipe_data.get('Name', path.stem), recipe_data))
                    ok(f"Added recipe: {recipe_data.get('Name', path.stem)}")
                    input(f"\n{C.DIM}Press Enter...{C.RESET}")
            elif choice == 'e':
                # Import from Excel/CSV
                recipes = await self._import_excel_for_batch()
                if recipes:
                    for recipe in recipes:
                        batch_queue.append(('generated_recipe', recipe.get('name', 'Excel_Recipe'), recipe))
                    ok(f"Added {len(recipes)} generated recipes from Excel")
                    input(f"\n{C.DIM}Press Enter...{C.RESET}")
            elif choice == 'c':
                batch_queue = []
                ok("Queue cleared")
                input(f"\n{C.DIM}Press Enter...{C.RESET}")
            elif choice == 'x' and batch_queue:
                await self._execute_batch(batch_queue)
                batch_queue = []  # Clear after execution
    
    async def _select_workflow_for_batch(self) -> tuple:
        """Select a workflow to add to batch queue."""
        clear()
        print_header("SELECT WORKFLOW")
        
        workflow_dir = BASE_DIR / "Library" / "Workflows"
        if not workflow_dir.exists():
            err("No workflows directory found")
            input(f"\n{C.DIM}Press Enter...{C.RESET}")
            return None
        
        # List workflows
        workflows = sorted(workflow_dir.glob("*.workflow.json"))
        
        if not workflows:
            err("No workflows found")
            input(f"\n{C.DIM}Press Enter...{C.RESET}")
            return None
        
        print(f"{C.BOLD}Available Workflows:{C.RESET}\n")
        for i, wf in enumerate(workflows, 1):
            try:
                with open(wf) as f:
                    data = json.load(f)
                name = data.get("WorkflowName", wf.stem)
                steps = len(data.get("Steps", []))
                print(f"  {C.CYAN}{i:2}{C.RESET}. {name} ({steps} steps)")
            except:
                print(f"  {C.CYAN}{i:2}{C.RESET}. {wf.stem} {C.RED}(invalid){C.RESET}")
        
        print(f"\n  {C.CYAN}0{C.RESET}. Cancel")
        print()
        
        choice = input(f"{C.BOLD}Select: {C.RESET}").strip()
        
        if choice == '0' or not choice.isdigit():
            return None
        
        idx = int(choice) - 1
        if 0 <= idx < len(workflows):
            wf_path = workflows[idx]
            with open(wf_path) as f:
                wf_data = json.load(f)
            return (wf_path, wf_data)
        
        return None
    
    async def _select_recipe_for_batch(self) -> tuple:
        """Select a recipe to add to batch queue."""
        clear()
        print_header("SELECT RECIPE")
        
        recipe_dir = BASE_DIR / "Library" / "Recipes"
        if not recipe_dir.exists():
            err("No recipes directory found")
            input(f"\n{C.DIM}Press Enter...{C.RESET}")
            return None
        
        # List recipes
        recipes = sorted(recipe_dir.glob("*.json"))
        
        if not recipes:
            err("No recipes found")
            input(f"\n{C.DIM}Press Enter...{C.RESET}")
            return None
        
        print(f"{C.BOLD}Available Recipes:{C.RESET}\n")
        for i, rc in enumerate(recipes, 1):
            try:
                with open(rc) as f:
                    data = json.load(f)
                name = data.get("Name", rc.stem)
                steps = len(data.get("Steps", []))
                print(f"  {C.CYAN}{i:2}{C.RESET}. {name} ({steps} steps)")
            except:
                print(f"  {C.CYAN}{i:2}{C.RESET}. {rc.stem} {C.RED}(invalid){C.RESET}")
        
        print(f"\n  {C.CYAN}0{C.RESET}. Cancel")
        print()
        
        choice = input(f"{C.BOLD}Select: {C.RESET}").strip()
        
        if choice == '0' or not choice.isdigit():
            return None
        
        idx = int(choice) - 1
        if 0 <= idx < len(recipes):
            rc_path = recipes[idx]
            with open(rc_path) as f:
                rc_data = json.load(f)
            return (rc_path, rc_data)
        
        return None
    
    async def _import_excel_for_batch(self) -> list:
        """Import pipetting plan from Excel/CSV and generate recipes."""
        import csv
        
        clear()
        print_header("IMPORT EXCEL/CSV PIPETTING PLAN")
        
        print(f"{C.BOLD}Expected CSV format:{C.RESET}")
        print(f"  Well,Liquid,Volume")
        print(f"  A1,Buffer,100")
        print(f"  A2,Sample1,50")
        print(f"  ...")
        print()
        
        # Get file path
        file_path = input(f"{C.BOLD}Enter file path (CSV or Excel): {C.RESET}").strip()
        
        if not file_path:
            return None
        
        # Handle quotes
        file_path = file_path.strip('"').strip("'")
        path = Path(file_path)
        
        if not path.exists():
            err(f"File not found: {file_path}")
            input(f"\n{C.DIM}Press Enter...{C.RESET}")
            return None
        
        # Parse operations from file
        operations = []
        
        try:
            if path.suffix.lower() == '.csv':
                with open(path, 'r', encoding='utf-8-sig') as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        # Flexible column names
                        well = row.get('Well') or row.get('well') or row.get('Destination') or row.get('dest') or ''
                        liquid = row.get('Liquid') or row.get('Source') or row.get('source') or row.get('liquid') or ''
                        volume = row.get('Volume') or row.get('volume') or row.get('Volume_uL') or row.get('vol') or '0'
                        source_well = row.get('Source_Well') or row.get('source_well') or ''
                        
                        if well:
                            try:
                                vol = float(str(volume).replace(',', '.').replace('µL', '').replace('ul', '').strip())
                            except:
                                vol = 0
                            
                            operations.append({
                                'well': well.upper().strip(),
                                'liquid': liquid.strip(),
                                'volume': vol,
                                'source_well': source_well.upper().strip() if source_well else ''
                            })
            
            elif path.suffix.lower() in ['.xlsx', '.xls']:
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(path, data_only=True)
                    ws = wb.active
                    
                    # Get headers from first row
                    headers = [str(cell.value or '').lower() for cell in ws[1]]
                    
                    # Find relevant columns
                    well_col = None
                    liquid_col = None
                    volume_col = None
                    source_col = None
                    
                    for i, h in enumerate(headers):
                        if h in ['well', 'destination', 'dest', 'target']:
                            well_col = i
                        elif h in ['liquid', 'source', 'reagent', 'sample']:
                            liquid_col = i
                        elif h in ['volume', 'vol', 'volume_ul', 'amount']:
                            volume_col = i
                        elif h in ['source_well', 'sourcewell', 'from']:
                            source_col = i
                    
                    if well_col is None or volume_col is None:
                        err("Could not find 'Well' and 'Volume' columns in Excel file")
                        input(f"\n{C.DIM}Press Enter...{C.RESET}")
                        return None
                    
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        well = str(row[well_col] or '') if len(row) > well_col else ''
                        liquid = str(row[liquid_col] or '') if liquid_col is not None and len(row) > liquid_col else ''
                        volume = row[volume_col] if len(row) > volume_col else 0
                        source_well = str(row[source_col] or '') if source_col is not None and len(row) > source_col else ''
                        
                        if well:
                            try:
                                vol = float(str(volume).replace(',', '.').replace('µL', '').replace('ul', '').strip())
                            except:
                                vol = 0
                            
                            operations.append({
                                'well': well.upper().strip(),
                                'liquid': liquid.strip(),
                                'volume': vol,
                                'source_well': source_well.upper().strip() if source_well else ''
                            })
                    
                except ImportError:
                    err("openpyxl not installed. Use CSV format or: pip install openpyxl")
                    input(f"\n{C.DIM}Press Enter...{C.RESET}")
                    return None
            else:
                err(f"Unsupported file format: {path.suffix}. Use .csv or .xlsx")
                input(f"\n{C.DIM}Press Enter...{C.RESET}")
                return None
                
        except Exception as e:
            err(f"Error parsing file: {e}")
            input(f"\n{C.DIM}Press Enter...{C.RESET}")
            return None
        
        if not operations:
            err("No valid operations found in file")
            input(f"\n{C.DIM}Press Enter...{C.RESET}")
            return None
        
        ok(f"Found {len(operations)} operations")
        print()
        
        # Plate type selection
        print(f"{C.BOLD}Plate capacity:{C.RESET}")
        print(f"  1. 96-well (default)")
        print(f"  2. 384-well")
        print(f"  3. 24-well")
        print(f"  4. 48-well")
        
        plate_choice = input(f"\n{C.BOLD}Select [1]: {C.RESET}").strip() or '1'
        plate_capacities = {'1': 96, '2': 384, '3': 24, '4': 48}
        plate_capacity = plate_capacities.get(plate_choice, 96)
        
        # Plate ID prefix
        plate_prefix = input(f"{C.BOLD}Plate ID prefix [PLATE]: {C.RESET}").strip() or "PLATE"
        
        # Ask for base protocol
        print()
        print(f"{C.BOLD}Select base protocol (for HAL config):{C.RESET}")
        recipe_dir = BASE_DIR / "Library" / "Recipes"
        recipe_files = sorted(recipe_dir.glob("*.json")) if recipe_dir.exists() else []
        
        if recipe_files:
            for i, rf in enumerate(recipe_files, 1):
                print(f"  {C.CYAN}{i}{C.RESET}. {rf.stem}")
        print(f"  {C.DIM}0. No base protocol{C.RESET}")
        
        base_protocol = None
        base_protocol_name = None
        if recipe_files:
            proto_choice = input(f"\n{C.BOLD}Select [0]: {C.RESET}").strip() or '0'
            if proto_choice.isdigit() and 1 <= int(proto_choice) <= len(recipe_files):
                with open(recipe_files[int(proto_choice) - 1]) as f:
                    base_protocol = json.load(f)
                base_protocol_name = recipe_files[int(proto_choice) - 1].stem
                ok(f"Using base protocol: {base_protocol_name}")
        
        # HAL config selection (if no base protocol)
        hal_config = None
        if not base_protocol:
            print()
            print(f"{C.BOLD}Select HAL config:{C.RESET}")
            hal_dir = BASE_DIR / "Library" / "HardwareConfig"
            hal_files = sorted(hal_dir.glob("*.json")) if hal_dir.exists() else []
            
            if hal_files:
                for i, hf in enumerate(hal_files, 1):
                    print(f"  {C.CYAN}{i}{C.RESET}. {hf.stem}")
                print(f"  {C.DIM}0. Default{C.RESET}")
                
                hal_choice = input(f"\n{C.BOLD}Select [0]: {C.RESET}").strip() or '0'
                if hal_choice.isdigit() and 1 <= int(hal_choice) <= len(hal_files):
                    hal_config = hal_files[int(hal_choice) - 1].name
        else:
            # Use HAL from base protocol
            hal_config = base_protocol.get("Requirements", {}).get("HAL")
        
        # Generate recipes from operations
        from datetime import datetime
        timestamp = datetime.now().strftime('%Y%m%d')
        
        total_wells = len(operations)
        num_recipes = (total_wells + plate_capacity - 1) // plate_capacity
        
        print()
        info(f"Will generate {num_recipes} recipe(s) for {total_wells} wells")
        print()
        
        confirm = input(f"{C.BOLD}Generate recipes? [Y/n]: {C.RESET}").strip().lower()
        if confirm == 'n':
            return None
        
        # Generate recipes
        recipes = []
        for recipe_idx in range(num_recipes):
            plate_id = f"{plate_prefix}-{timestamp}-{str(recipe_idx + 1).zfill(3)}"
            start = recipe_idx * plate_capacity
            end = min((recipe_idx + 1) * plate_capacity, total_wells)
            plate_ops = operations[start:end]
            
            # Tag operations with plate ID
            for op in plate_ops:
                op['plate_id'] = plate_id
            
            # Group operations by liquid for efficiency
            ops_by_liquid = {}
            for op in plate_ops:
                liquid = op.get('liquid') or 'default'
                if liquid not in ops_by_liquid:
                    ops_by_liquid[liquid] = []
                ops_by_liquid[liquid].append(op)
            
            # Generate recipe steps
            steps = []
            for liquid, ops in ops_by_liquid.items():
                transfers = []
                for op in ops:
                    transfers.append({
                        'dest_well': op['well'],
                        'volume_ul': op['volume'],
                        'source_well': op.get('source_well') or 'A1',
                        'plate_id': op.get('plate_id')
                    })
                
                steps.append({
                    'command': 'TransferLiquid',
                    'liquid': liquid,
                    'transfers': transfers
                })
            
            recipe = {
                'name': f"Excel_{plate_prefix}_{plate_id}",
                'description': f"Auto-generated from Excel. Plate: {plate_id}",
                'plate_ids': [plate_id],
                'hal_config': hal_config or "Standard_Flex_Setup",
                'base_protocol': base_protocol_name,
                'steps': steps,
                'metadata': {
                    'generated': datetime.now().isoformat(),
                    'total_operations': len(plate_ops),
                    'source_file': str(path.name)
                }
            }
            recipes.append(recipe)
        
        ok(f"Generated {len(recipes)} recipes")
        return recipes
        
        return None
    
    async def _execute_batch(self, queue: list):
        """Execute all items in the batch queue."""
        clear()
        print_header("EXECUTING BATCH")
        
        total = len(queue)
        results = []  # List of (name, success, error)
        
        print(f"{C.BOLD}Starting batch of {total} item(s)...{C.RESET}\n")
        
        # Select HAL config once for all items (if not already set)
        hal_config = None
        
        # Tecan analysis options
        run_tecan_analysis = False
        tecan_protocol = None
        tecan_id = None
        
        # Check if any item needs Opentrons
        needs_opentrons = any(
            (item_type == 'workflow' and any("opentrons" in s.get("Instrument", "").lower() for s in data.get("Steps", [])))
            or item_type in ('recipe', 'generated_recipe')
            for item_type, _, data in queue
        )
        
        if needs_opentrons:
            # Check if generated recipes already have HAL config
            generated_hal = None
            for item_type, _, data in queue:
                if item_type == 'generated_recipe' and data.get('hal_config'):
                    generated_hal = data.get('hal_config')
                    break
            
            if generated_hal:
                hal_config = generated_hal
                info(f"Using HAL from generated recipes: {hal_config}")
            else:
                print(f"{C.BOLD}Select HAL config for Opentrons:{C.RESET}\n")
                
                hal_dir = BASE_DIR / "Library" / "HardwareConfig"
                hal_files = []
                if hal_dir.exists():
                    hal_files = sorted(hal_dir.glob("*.json"))
                
                if hal_files:
                    for i, hf in enumerate(hal_files, 1):
                        print(f"  {C.CYAN}{i}{C.RESET}. {hf.stem}")
                    print(f"  {C.DIM}0. Skip HAL{C.RESET}")
                    
                    choice = input(f"\n{C.BOLD}Select [1-{len(hal_files)}]: {C.RESET}").strip()
                    if choice.isdigit() and 1 <= int(choice) <= len(hal_files):
                        hal_config = hal_files[int(choice) - 1].name
                        ok(f"Using HAL: {hal_config}")
            
            print()
        
        # Option for Tecan analysis after pipetting
        has_generated_recipes = any(t == 'generated_recipe' for t, _, _ in queue)
        if has_generated_recipes:
            analysis_dir = BASE_DIR / "Library" / "Analysis"
            tecan_protocols = []
            if analysis_dir.exists():
                tecan_protocols = sorted(analysis_dir.glob("*.mdfx"))
            
            if tecan_protocols:
                tecan_choice = input(f"{C.BOLD}Run Tecan analysis after pipetting? [y/N]: {C.RESET}").strip().lower()
                if tecan_choice == 'y':
                    print(f"\n{C.BOLD}Select Tecan Protocol:{C.RESET}")
                    for i, tp in enumerate(tecan_protocols, 1):
                        print(f"  {C.CYAN}{i}{C.RESET}. {tp.stem}")
                    
                    tp_choice = input(f"\n{C.BOLD}Select [1-{len(tecan_protocols)}]: {C.RESET}").strip()
                    if tp_choice.isdigit() and 1 <= int(tp_choice) <= len(tecan_protocols):
                        tecan_protocol = tecan_protocols[int(tp_choice) - 1].name
                        run_tecan_analysis = True
                        ok(f"Will run Tecan analysis with: {tecan_protocol}")
                    print()
        
        # Ensure LabCore is ready
        await self.lab_core.discover()
        
        # Find Opentrons instrument ID
        opentrons_id = None
        for instr in self.lab_core.instruments:
            if "opentrons" in instr.type.lower() or "opentrons" in instr.name.lower():
                opentrons_id = instr.id
                break
        
        # Find Tecan instrument for analysis
        if run_tecan_analysis:
            for instr in self.lab_core.instruments:
                if "tecan" in instr.type.lower() or "tecan" in instr.name.lower():
                    tecan_id = instr.id
                    break
            if not tecan_id:
                warn("Tecan not found - analysis will be skipped")
                run_tecan_analysis = False
        
        # Execute each item
        for idx, (item_type, name, data) in enumerate(queue, 1):
            print(f"{C.BOLD}{'═' * 60}{C.RESET}")
            print(f"{C.BOLD}[{idx}/{total}] {name}{C.RESET}")
            print(f"{C.DIM}Type: {item_type}{C.RESET}")
            print(f"{C.BOLD}{'═' * 60}{C.RESET}\n")
            
            success = True
            error_msg = ""
            
            try:
                if item_type == 'workflow':
                    # Execute workflow steps
                    steps = data.get("Steps", [])
                    
                    for step in steps:
                        step_num = step.get("StepNumber", "?")
                        instrument = step.get("Instrument", "")
                        action = step.get("Action", "")
                        params = step.get("Parameters", {}).copy()
                        
                        print(f"  [{step_num}] {instrument} -> {action}", end=" ", flush=True)
                        
                        # Handle Manual Station
                        if "manual" in instrument.lower():
                            manual_ok = await self._handle_manual_step(step)
                            if manual_ok:
                                ok("")
                            else:
                                err("Cancelled")
                                success = False
                                error_msg = "Manual step cancelled"
                                break
                            continue
                        
                        # Inject HAL config for Opentrons
                        if "opentrons" in instrument.lower() and hal_config:
                            params["HALConfig"] = hal_config
                            params["UseHAL"] = True
                        
                        # Execute
                        instrument_id = instrument.lower().replace(" ", "_")
                        result = await self.lab_core.execute_command(instrument_id, action, params)
                        
                        if not result.success:
                            err(f"FAILED: {result.error}")
                            success = False
                            error_msg = result.error
                            break
                        elif result.data and isinstance(result.data, dict):
                            if result.data.get("status", "").lower() == "error":
                                err(f"FAILED: {result.data.get('message', 'Error')}")
                                success = False
                                error_msg = result.data.get("message", "Error")
                                break
                        
                        ok("")
                
                elif item_type == 'recipe':
                    # Execute recipe via RunProtocol
                    if opentrons_id:
                        info(f"Executing recipe on Opentrons...")
                        
                        # Add HAL config if needed
                        if hal_config and 'Requirements' not in data:
                            data['Requirements'] = {'HAL': hal_config}
                        elif hal_config:
                            data['Requirements']['HAL'] = hal_config
                        
                        result = await self.lab_core.execute_command(
                            opentrons_id,
                            "RunProtocol",
                            {"protocol_content": json.dumps(data), "protocol_type": "json"}
                        )
                        
                        if not result.success:
                            success = False
                            error_msg = result.error
                            err(f"FAILED: {error_msg}")
                        else:
                            ok("Recipe completed")
                    else:
                        err("Opentrons not found")
                        success = False
                        error_msg = "Opentrons instrument not discovered"
                
                elif item_type == 'generated_recipe':
                    # Execute generated recipe from Excel
                    plate_id = data.get('plate_ids', ['?'])[0]
                    info(f"Executing generated recipe for plate: {plate_id}")
                    
                    if opentrons_id:
                        # Convert to Opentrons format
                        opentrons_recipe = self._convert_generated_to_opentrons(data, hal_config)
                        
                        step_count = len(data.get('steps', []))
                        print(f"  Running {step_count} step(s)...")
                        
                        result = await self.lab_core.execute_command(
                            opentrons_id,
                            "RunProtocol",
                            {"protocol_content": json.dumps(opentrons_recipe), "protocol_type": "json"}
                        )
                        
                        if not result.success:
                            success = False
                            error_msg = result.error
                            err(f"FAILED: {error_msg}")
                        else:
                            ok(f"Plate {plate_id} completed")
                            
                            # Run Tecan analysis if enabled
                            if run_tecan_analysis and tecan_id and tecan_protocol:
                                info(f"Running Tecan analysis for plate {plate_id}...")
                                try:
                                    from datetime import datetime
                                    tecan_result = await self.lab_core.execute_command(
                                        tecan_id,
                                        "RunMeasurement",
                                        {
                                            "protocol_file": tecan_protocol,
                                            "plate_id": plate_id,
                                            "sample_set_id": f"batch_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
                                        }
                                    )
                                    if tecan_result.success:
                                        ok(f"  Tecan analysis completed for plate {plate_id}")
                                    else:
                                        warn(f"  Tecan analysis failed: {tecan_result.error}")
                                except Exception as tecan_err:
                                    warn(f"  Tecan analysis error: {tecan_err}")
                    else:
                        # Simulation mode
                        warn("Opentrons not found - simulating execution")
                        steps = data.get('steps', [])
                        for step in steps:
                            cmd = step.get('command', 'Transfer')
                            transfers = step.get('transfers', [])
                            print(f"    [SIM] {cmd}: {len(transfers)} operations")
                            await asyncio.sleep(0.3)
                        ok(f"Plate {plate_id} simulated")
                        
                        # Simulate Tecan analysis
                        if run_tecan_analysis and tecan_protocol:
                            info(f"[SIM] Tecan analysis for plate {plate_id} with {tecan_protocol}")
                
            except Exception as e:
                success = False
                error_msg = str(e)
                err(f"Exception: {error_msg}")
            
            results.append((name, success, error_msg))
            
            if not success:
                # Ask whether to continue batch
                cont = input(f"\n{C.BOLD}Item failed. Continue batch? [y/N]: {C.RESET}").strip().lower()
                if cont != 'y':
                    break
            
            print()
        
        # Summary
        print(f"\n{C.BOLD}{'═' * 60}{C.RESET}")
        print(f"{C.BOLD}BATCH COMPLETE{C.RESET}")
        print(f"{C.BOLD}{'═' * 60}{C.RESET}\n")
        
        success_count = sum(1 for _, s, _ in results if s)
        fail_count = len(results) - success_count
        
        print(f"  {C.GREEN}Succeeded: {success_count}{C.RESET}")
        print(f"  {C.RED}Failed: {fail_count}{C.RESET}")
        
        if fail_count > 0:
            print(f"\n{C.BOLD}Failed items:{C.RESET}")
            for item_name, success, error in results:
                if not success:
                    print(f"  {C.RED}✗{C.RESET} {item_name}: {error}")
        
        input(f"\n{C.DIM}Press Enter...{C.RESET}")
    
    def _convert_generated_to_opentrons(self, recipe_data: dict, hal_config: str = None) -> dict:
        """Convert a generated pipetting recipe to Opentrons JSON format."""
        steps = recipe_data.get('steps', [])
        opentrons_steps = []
        
        for step in steps:
            cmd = step.get('command', 'TransferLiquid')
            transfers = step.get('transfers', [])
            liquid = step.get('liquid', 'default')
            
            if cmd == 'TransferLiquid' and transfers:
                for transfer in transfers:
                    opentrons_steps.append({
                        "Command": "Transfer",
                        "SourceLabware": transfer.get('source_labware', 'reservoir_1'),
                        "SourceWell": transfer.get('source_well', 'A1'),
                        "DestLabware": transfer.get('dest_labware', 'plate_1'),
                        "DestWell": transfer.get('dest_well'),
                        "Volume": transfer.get('volume_ul', 0),
                        "Comment": f"{liquid} to {transfer.get('dest_well')}"
                    })
            else:
                opentrons_steps.append({
                    "Command": cmd,
                    **{k: v for k, v in step.items() if k not in ['command', 'transfers']}
                })
        
        return {
            "Name": recipe_data.get('name', 'Generated_Recipe'),
            "Description": recipe_data.get('description', 'Auto-generated from Excel'),
            "Requirements": {
                "HAL": hal_config or recipe_data.get('hal_config') or "Standard_Flex_Setup"
            },
            "Steps": opentrons_steps,
            "Metadata": recipe_data.get('metadata', {})
        }
    
    #                         STATUS MENU
    
    async def status_menu(self):
        """System status overview."""
        clear()
        print_header("SYSTEM STATUS")
        
        servers = self.registry.list_servers()
        
        print(f"{C.BOLD}Servers ({len(servers)}):{C.RESET}\n")
        
        for server in servers:
            # Status icons
            server_icon = f"{C.GREEN}[Y]{C.RESET}" if server.server_online else f"{C.RED}[N]{C.RESET}"
            hw_icon = f"{C.GREEN}[Y]{C.RESET}" if server.hardware_online else f"{C.RED}[N]{C.RESET}"
            
            print(f"  {server.name}")
            print(f"    Server:   {server_icon} ({server.address})")
            print(f"    Hardware: {hw_icon} ({server.hardware_status})")
            print(f"    Type:     {server.server_type or 'unknown'}")
            print(f"    Commands: {len(server.get_all_commands())}")
            print()
        
        # Refresh option
        print(f"  {C.CYAN}r{C.RESET}. Refresh health status")
        print(f"  {C.CYAN}0{C.RESET}. <- Back")
        
        choice = input(f"\n{C.BOLD}Select: {C.RESET}").strip().lower()
        
        if choice == 'r':
            info("Checking health...")
            for server in servers:
                await self.registry.client.check_health(server)
            ok("Health check complete")
            await asyncio.sleep(1)
            await self.status_menu()


#                              MAIN

async def main():
    """Entry point."""
    console = PnPConsole()
    await console.run()


def main_sync():
    """Synchronous entry point."""
    try:
        asyncio.run(main())
    except KeyboardInterrupt:
        print(f"\n{C.GREEN}Goodbye!{C.RESET}")


if __name__ == "__main__":
    main_sync()
