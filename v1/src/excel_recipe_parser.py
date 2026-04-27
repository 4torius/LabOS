"""
Excel Recipe Parser for Chemical Pipetting Plans

Flexible parser that:
1. Automatically detects reagent columns from Excel headers
2. Allows user configuration of column grouping and delays
3. Generates recipes with embedded Opentrons Delay commands

No hardcoded reagent names - everything is read from Excel.
"""

import io
import json
import re
from datetime import datetime
from typing import List, Dict, Any, Optional, Tuple
from pathlib import Path

# Plate catalog helpers

_PLATE_CACHE: Dict[str, Dict] = {}

def _load_plate_catalog(library_dir: Optional[Path] = None) -> Dict[str, Dict]:
    """Load all plate definitions from Library/Labware/Plates/. Cached after first call."""
    global _PLATE_CACHE
    if _PLATE_CACHE:
        return _PLATE_CACHE
    if library_dir is None:
        library_dir = Path(__file__).parent.parent / "Library"
    plates_dir = library_dir / "Labware" / "Plates"
    if not plates_dir.exists():
        return {}
    for f in plates_dir.glob("*.plate.json"):
        try:
            data = json.loads(f.read_text(encoding='utf-8'))
            _PLATE_CACHE[data["id"]] = data
        except Exception:
            pass
    return _PLATE_CACHE


def get_plate_info(plate_id: str, library_dir: Optional[Path] = None) -> Optional[Dict]:
    """Return plate definition dict for the given Opentrons load name, or None."""
    return _load_plate_catalog(library_dir).get(plate_id)


def get_plate_well_regex(plate_id: str, library_dir: Optional[Path] = None) -> str:
    """Return a regex pattern that matches valid well IDs for the given plate.

    Falls back to 96-well (A-H) if the plate is unknown.
    """
    plate = get_plate_info(plate_id, library_dir)
    if plate:
        row_labels = plate.get("row_labels", ["A","B","C","D","E","F","G","H"])
        max_col = plate.get("columns", 12)
        if len(row_labels) <= 8:
            row_pat = "A-H"
        elif len(row_labels) <= 16:
            row_pat = "A-P"
        else:
            row_pat = "A-Z"
        return rf'^[{row_pat}]\d{{1,{"2" if max_col <= 12 else "2"}}}$'
    return r'^[A-P]\d{1,2}$'  # permissive default: accept up to 384-well


def generate_all_wells(plate_id: str, library_dir: Optional[Path] = None) -> List[str]:
    """Return ordered list of all well IDs for the plate (column-major: A1,B1,...,H1,A2,...).

    Useful when the recipe needs to fill every well regardless of Excel input.
    """
    plate = get_plate_info(plate_id, library_dir)
    if plate is None:
        # Default 96-well
        rows, cols = ["A","B","C","D","E","F","G","H"], 12
    else:
        rows = plate.get("row_labels", ["A","B","C","D","E","F","G","H"])
        cols = plate.get("columns", 12)
    wells = []
    for c in range(1, cols + 1):
        for r in rows:
            wells.append(f"{r}{c}")
    return wells


def parse_excel_columns(file_content: bytes, filename: str = "recipe.xlsx") -> Dict[str, Any]:
    """
    Parse an Excel file and extract column information.
    
    This function ONLY detects columns - it does NOT make decisions about
    phases or delays. The user will configure that separately.
    
    Returns:
        {
            'filename': str,
            'sheets': [
                {
                    'sheet_name': str,
                    'columns': [
                        {'index': int, 'name': str, 'header': str, 'type': 'well'|'reagent'|'other'},
                        ...
                    ],
                    'wells': [{'well': 'A1', 'reagents': {'HA': 10.0, 'DMTMM': 5.0}}],
                    'preview_rows': int
                }
            ],
            'summary': {...}
        }
    """
    try:
        import openpyxl
    except ImportError:
        return {'error': 'openpyxl not installed. Run: pip install openpyxl'}
    
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
    except Exception as e:
        return {'error': f'Failed to load Excel file: {str(e)}'}
    
    sheets_data = []
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_result = _parse_sheet_columns(ws, sheet_name)
        sheets_data.append(sheet_result)
    
    return {
        'filename': filename,
        'sheets': sheets_data,
        'summary': {
            'total_sheets': len(wb.sheetnames),
            'sheet_names': wb.sheetnames
        }
    }


def _parse_sheet_columns(ws, sheet_name: str) -> Dict[str, Any]:
    """Parse a single worksheet to extract column information."""
    
    # Find the header row (contains volume columns with µL or uL)
    header_row = None
    headers = []
    
    for row_idx in range(1, min(20, ws.max_row + 1)):
        row_values = [str(cell.value or '').strip() for cell in ws[row_idx]]
        
        # Look for "Well" column AND volume columns
        has_well = any('well' in v.lower() for v in row_values)
        has_volume = any('ul' in v.lower() or 'µl' in v.lower() for v in row_values)
        
        if has_well and has_volume:
            header_row = row_idx
            headers = row_values
            break
    
    if header_row is None:
        return {
            'sheet_name': sheet_name,
            'error': 'Could not find header row with Well and volume columns'
        }
    
    # Parse all columns
    columns = []
    well_col = None
    reagent_columns = []
    
    for col_idx, header in enumerate(headers):
        h_lower = header.lower()
        
        col_info = {
            'index': col_idx,
            'header': header,
            'name': header,
            'type': 'other'
        }
        
        if 'well' in h_lower and well_col is None:
            well_col = col_idx
            col_info['type'] = 'well'
            col_info['name'] = 'Well'
        elif '(ul)' in h_lower or '(µl)' in h_lower or 'ul)' in h_lower:
            # Extract clean reagent name
            reagent_name = re.sub(r'\s*\(.*\).*', '', header).strip()
            reagent_name = re.sub(r'\s*\d+[%mM]*\s*$', '', reagent_name).strip()
            
            # Skip "Total Volume" - it's calculated
            if 'total' in reagent_name.lower() and 'volume' in reagent_name.lower():
                col_info['type'] = 'calculated'
            else:
                col_info['type'] = 'reagent'
                col_info['name'] = reagent_name
                reagent_columns.append(col_idx)
        elif 'ratio' in h_lower or 'molar' in h_lower:
            col_info['type'] = 'ratio'
        
        columns.append(col_info)
    
    if well_col is None:
        return {
            'sheet_name': sheet_name,
            'columns': columns,
            'error': 'Could not find Well column'
        }
    
    if not reagent_columns:
        return {
            'sheet_name': sheet_name,
            'columns': columns,
            'error': 'Could not find reagent volume columns'
        }
    
    # Parse data rows
    wells_data = []
    reagent_names = [columns[i]['name'] for i in reagent_columns]
    
    for row_idx in range(header_row + 1, ws.max_row + 1):
        row = list(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
        
        if len(row) <= well_col:
            continue
            
        well = str(row[well_col] or '').strip().upper()
        
        # Accept well IDs up to 384-well (A-P, cols 1-24); permissive to support any plate
        if not well or not re.match(r'^[A-P]\d{1,2}$', well):
            continue
        
        well_entry = {
            'well': well,
            'reagents': {}
        }
        
        # Get volumes for each reagent
        for reagent_col_idx in reagent_columns:
            reagent_name = columns[reagent_col_idx]['name']
            if len(row) > reagent_col_idx:
                try:
                    vol = float(row[reagent_col_idx] or 0)
                    if vol > 0:
                        well_entry['reagents'][reagent_name] = vol
                except:
                    pass
        
        if well_entry['reagents']:
            wells_data.append(well_entry)
    
    return {
        'sheet_name': sheet_name,
        'header_row': header_row,
        'columns': columns,
        'reagent_columns': [columns[i] for i in reagent_columns],
        'wells': wells_data,
        'summary': {
            'total_wells': len(wells_data),
            'reagent_count': len(reagent_columns),
            'reagents': reagent_names
        }
    }


def generate_recipe_with_config(
    sheet_data: Dict,
    config: Dict,
    recipe_name: Optional[str] = None
) -> Dict:
    """
    Generate a recipe from parsed sheet data using user-provided configuration.
    
    The config specifies how columns are grouped and delays between groups:
    {
        "groups": [
            {
                "name": "Phase A",
                "columns": ["HA", "DMTMM"],
                "delay_after_seconds": 1800  # 30 minutes
            },
            {
                "name": "Phase B",
                "columns": ["AMF"],
                "delay_after_seconds": 0
            }
        ],
        "target_plate": "corning_96_wellplate_360ul_flat",
        "reservoir_type": "nest_12_reservoir_15ml",
        "pipette": "left"
    }
    
    This generates a SINGLE recipe with embedded Delay commands.
    """
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    
    if not recipe_name:
        recipe_name = re.sub(r'[^\w]', '_', sheet_data.get('sheet_name', 'Recipe'))
    
    groups = config.get('groups', [])
    wells = sheet_data.get('wells', [])
    
    if not wells:
        return {'error': 'No well data found'}
    
    if not groups:
        # Default: all reagents in one group, no delays
        all_reagents = sheet_data.get('summary', {}).get('reagents', [])
        groups = [{'name': 'All', 'columns': all_reagents, 'delay_after_seconds': 0}]
    
    # Build Requirements section with GENERIC numbered sources (Source_1, Source_2, ...)
    # This allows reusing the same HAL for different recipes
    requirements = {}
    all_reagents_in_config = []
    reagent_to_source_map = {}  # Maps reagent name -> generic source (e.g., "HA" -> "Source_1")
    source_counter = 1
    
    for group in groups:
        for reagent in group.get('columns', []):
            source_key = f"Source_{source_counter}"
            requirements[source_key] = source_key  # Generic numbered source
            reagent_to_source_map[reagent] = source_key
            all_reagents_in_config.append(reagent)
            source_counter += 1
    
    requirements["TargetPlate"] = "TargetPlate"
    requirements["MyTips"] = "MyTips"
    requirements["MyTrash"] = "MyTrash"
    
    # Build Steps with delays between groups
    steps = []
    
    for group_idx, group in enumerate(groups):
        group_name = group.get('name', f'Group_{group_idx+1}')
        columns = group.get('columns', [])
        delay_seconds = group.get('delay_after_seconds', 0)
        
        if not columns:
            continue
        
        # Add group comment
        steps.append({
            "Command": "Comment",
            "Text": f"=== {group_name}: {', '.join(columns)} ==="
        })
        
        # Process each reagent in this group
        for reagent in columns:
            # Use the generic source from the mapping
            source_name = reagent_to_source_map.get(reagent, f"Source_{reagent}")
            
            # Get LiquidClass for this reagent (if specified)
            # Config can have: reagent_liquid_classes: {"HA": "Aqueous", "DMTMM": "Viscous", ...}
            reagent_liquid_classes = config.get('reagent_liquid_classes', {})
            liquid_class = reagent_liquid_classes.get(reagent, None)
            
            # Collect transfers for this reagent
            transfers = []
            for well_data in wells:
                vol = well_data['reagents'].get(reagent, 0)
                if vol > 0:
                    transfers.append({
                        'well': well_data['well'],
                        'volume': round(vol, 2)
                    })
            
            if not transfers:
                continue
            
            # PickUp Tip
            steps.append({
                "Command": "PickUpTip",
                "PipetteMount": config.get('pipette', 'left'),
                "TipRack": "MyTips"
            })
            
            # Build volumes and destinations
            volumes_list = [t['volume'] for t in transfers]
            dest_wells = [t['well'] for t in transfers]
            unique_volumes = set(volumes_list)
            
            # Always use Distribute - it's the most efficient
            # If all volumes are the same, use single volume; otherwise use list
            if len(unique_volumes) == 1:
                # Same volume for all wells
                volume_param = volumes_list[0]
                description = f"Distribute {volume_param}µL {reagent} to {len(dest_wells)} wells"
            else:
                # Variable volumes per well
                volume_param = volumes_list
                description = f"Distribute {min(volumes_list)}-{max(volumes_list)}µL {reagent} to {len(dest_wells)} wells"
            
            # Build distribute command
            distribute_cmd = {
                "Command": "Distribute",
                "PipetteMount": config.get('pipette', 'left'),
                "Volume": volume_param,
                "Source": f"{source_name}:A1",
                "Destinations": [f"TargetPlate:{w}" for w in dest_wells],
                "NewTip": "never",
                "DisposalVolume": 0,
                "Description": description
            }
            
            # Add LiquidClass if specified for this reagent
            if liquid_class:
                distribute_cmd["LiquidClass"] = liquid_class
            
            steps.append(distribute_cmd)
            
            # Drop Tip
            steps.append({
                "Command": "DropTip",
                "PipetteMount": config.get('pipette', 'left'),
                "TrashLocation": "MyTrash"
            })
        
        # Add delay AFTER this group (if delay > 0)
        if delay_seconds > 0:
            delay_minutes = delay_seconds / 60
            steps.append({
                "Command": "Delay",
                "Seconds": delay_seconds,
                "Description": f"Wait {delay_minutes:.1f} minutes after {group_name}"
            })
    
    # Check if any delays were added
    has_delays = any(g.get('delay_after_seconds', 0) > 0 for g in groups)
    
    # Build recipe
    recipe = {
        'ProtocolName': recipe_name,
        'Description': f"Generated from Excel. Groups: {', '.join([g.get('name', '?') for g in groups])}",
        'Requirements': requirements,
        'Steps': steps,
        'Metadata': {
            'Generated': timestamp,
            'Source': 'excel_import_configurable',
            'TotalWells': len(wells),
            'PlateId': config.get('target_plate', ''),
            'PlateRows': config.get('plate_rows', 8),
            'PlateColumns': config.get('plate_columns', 12),
            'PlateTotalWells': config.get('plate_total_wells', 96),
            'Reagents': all_reagents_in_config,
            'Groups': [{'name': g.get('name'), 'columns': g.get('columns'), 'delay_seconds': g.get('delay_after_seconds', 0)} for g in groups],
            'ReagentMapping': reagent_to_source_map,
            'HasDelays': has_delays
        },
        'Configuration': config
    }
    
    return recipe


def generate_default_config(sheet_data: Dict) -> Dict:
    """
    Generate a default configuration with all reagents in one group (no delays).
    User can modify this in the UI.
    """
    reagents = sheet_data.get('summary', {}).get('reagents', [])
    
    return {
        'groups': [
            {
                'name': 'All Reagents',
                'columns': reagents,
                'delay_after_seconds': 0
            }
        ],
        'target_plate': 'corning_96_wellplate_360ul_flat',
        'reservoir_type': 'nest_12_reservoir_15ml',
        'pipette': 'left'
    }


def save_recipe_to_library(recipe: Dict, library_path: Path) -> str:
    """Save generated recipe to the Library/Recipes folder."""
    import json
    
    recipes_dir = library_path / "Recipes"
    recipes_dir.mkdir(parents=True, exist_ok=True)
    
    timestamp = recipe.get('Metadata', {}).get('Generated', datetime.now().strftime('%Y%m%d_%H%M%S'))
    filename = f"{recipe['ProtocolName']}_{timestamp}.json"
    filepath = recipes_dir / filename
    
    with open(filepath, 'w', encoding='utf-8') as f:
        json.dump(recipe, f, indent=2, ensure_ascii=False)
    
    return filename


def generate_hal_config_from_recipe(recipe: Dict, hal_name: Optional[str] = None) -> Dict:
    """
    Generate a GENERIC HAL configuration that can be reused across recipes.
    Uses numbered sources (Source_1, Source_2, ...) instead of reagent-specific names.
    The recipe's Metadata.ReagentMapping tells which reagent goes where.
    """
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    
    if not hal_name:
        hal_name = "Generic_HAL"
    
    # Count how many sources are needed
    num_sources = 0
    for key in recipe.get('Requirements', {}):
        if key.startswith('Source_'):
            num_sources += 1
    
    # Slot assignments for reservoirs
    # Priority: A2 → B2 → C2 → D3 → D1 → C1 → B1 → C3
    # A1=HeaterShaker, A3=Trash, D2=TargetPlate, B3=TipRack
    available_reservoir_slots = ['A2', 'B2', 'C2', 'D3', 'D1', 'C1', 'B1', 'C3']
    
    labware = {}
    slot_assignments = {}
    
    # Tip racks - Column 3: B3 primary, C3 backup
    # A1=HeaterShaker, A3=Trash
    labware['MyTips'] = {
        'LoadName': 'opentrons_flex_96_tiprack_1000ul',
        'Slot': 'B3',
        'DisplayName': 'Tip rack 1000µL'
    }
    slot_assignments['B3'] = 'MyTips'
    
    labware['MyTips_2'] = {
        'LoadName': 'opentrons_flex_96_tiprack_1000ul',
        'Slot': 'C3',
        'DisplayName': 'Tip rack 1000µL (backup)'
    }
    slot_assignments['C3'] = 'MyTips_2'
    
    # Target plate
    config = recipe.get('Configuration', {})
    target_plate_type = config.get('target_plate', 'corning_96_wellplate_360ul_flat')
    
    labware['TargetPlate'] = {
        'LoadName': target_plate_type,
        'Slot': 'D2',
        'DisplayName': 'Target plate for reactions'
    }
    slot_assignments['D2'] = 'TargetPlate'
    
    # Generic numbered reservoirs (Source_1, Source_2, ...)
    reservoir_type = config.get('reservoir_type', 'nest_12_reservoir_15ml')
    
    # Get reagent mapping if available (for display names)
    reagent_mapping = recipe.get('Metadata', {}).get('ReagentMapping', {})
    source_to_reagent = {v: k for k, v in reagent_mapping.items()}  # Reverse: Source_1 -> HA
    
    for i in range(num_sources):
        if i >= len(available_reservoir_slots):
            break
        
        slot = available_reservoir_slots[i]
        source_key = f'Source_{i+1}'
        reagent_hint = source_to_reagent.get(source_key, f'Reagent {i+1}')
        
        labware[source_key] = {
            'LoadName': reservoir_type,
            'Slot': slot,
            'DisplayName': f'{source_key} (e.g. {reagent_hint})',
            'Well': 'A1'
        }
        slot_assignments[slot] = source_key
    
    # Use ConfigName for compatibility with existing HAL format
    return {
        'ConfigName': hal_name,
        'Description': f'Generic HAL with {num_sources} source slots. Reusable for multiple recipes.',
        'Generated': timestamp,
        'Trash': {
            'MyTrash': {'Type': 'TrashBin', 'Slot': 'A3'}
        },
        'Labware': labware,
        'Pipettes': {
            'left': 'flex_1channel_1000'
        },
        'Metadata': {
            'source': 'excel_recipe_configurable',
            'total_sources': num_sources,
            'slot_assignments': slot_assignments,
            'note': 'This HAL uses generic source names. Check recipe metadata for reagent assignments.'
        }
    }


def generate_generic_hal(
    num_sources: int = 6,
    hal_name: str = "Generic_HAL",
    target_plate: str = "corning_96_wellplate_360ul_flat",
    reservoir_type: str = "nest_12_reservoir_15ml",
    pipette: str = "flex_1channel_1000"
) -> Dict:
    """
    Generate a completely generic HAL configuration with numbered source slots.
    
    This HAL can be reused for ANY recipe - just put the reagents in corresponding Source_N positions.
    
    Args:
        num_sources: Number of source reservoirs (Source_1 to Source_N)
        hal_name: Name for the HAL configuration
        target_plate: Labware type for target plate
        reservoir_type: Labware type for reagent reservoirs
        pipette: Pipette name for left mount
        
    Returns:
        HAL configuration dict compatible with opentrons_client.py
    """
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    
    # Available slots for reservoirs
    # Priority: A2 → B2 → C2 → D3 → D1 → C1 → B1 → C3
    available_reservoir_slots = ['A2', 'B2', 'C2', 'D3', 'D1', 'C1', 'B1', 'C3']
    max_sources = min(num_sources, len(available_reservoir_slots))
    
    labware = {}
    slot_assignments = {}
    
    # Tip racks - Column 3: B3 primary, C3 backup
    labware['MyTips'] = {
        'LoadName': 'opentrons_flex_96_tiprack_1000ul',
        'Slot': 'B3',
        'DisplayName': 'Tip rack 1000µL'
    }
    slot_assignments['B3'] = 'MyTips'
    
    labware['MyTips_2'] = {
        'LoadName': 'opentrons_flex_96_tiprack_1000ul',
        'Slot': 'C3',
        'DisplayName': 'Tip rack 1000µL (backup)'
    }
    slot_assignments['C3'] = 'MyTips_2'
    
    # Target plate
    labware['TargetPlate'] = {
        'LoadName': target_plate,
        'Slot': 'D2',
        'DisplayName': 'Target plate'
    }
    slot_assignments['D2'] = 'TargetPlate'
    
    # Generic numbered reservoirs
    for i in range(max_sources):
        slot = available_reservoir_slots[i]
        source_key = f'Source_{i+1}'
        
        labware[source_key] = {
            'LoadName': reservoir_type,
            'Slot': slot,
            'DisplayName': f'{source_key} - Reagent reservoir',
            'Well': 'A1'
        }
        slot_assignments[slot] = source_key
    
    return {
        'ConfigName': hal_name,
        'Description': f'Generic reusable HAL with {max_sources} source slots (Source_1 to Source_{max_sources}). Works with any recipe using numbered sources.',
        'Generated': timestamp,
        'Trash': {
            'MyTrash': {'Type': 'TrashBin', 'Slot': 'A3'}
        },
        'Labware': labware,
        'Pipettes': {
            'left': pipette
        },
        'Metadata': {
            'source': 'generic_hal_generator',
            'total_sources': max_sources,
            'slot_assignments': slot_assignments,
            'note': 'Load reagents into Source_1, Source_2, etc. Check recipe metadata for reagent order.'
        }
    }


def save_hal_config_to_library(hal_config: Dict, library_path: Path) -> str:
    """Save generated HAL config to the Library/HardwareConfig folder."""
    import json
    
    hal_dir = library_path / "HardwareConfig"
    hal_dir.mkdir(parents=True, exist_ok=True)
    
    timestamp = hal_config.get('Generated', datetime.now().strftime('%Y%m%d_%H%M%S'))
    config_name = hal_config.get('ConfigName') or hal_config.get('HALName', 'HAL_Config')
    filename = f"{config_name}_{timestamp}.json"
    filepath = hal_dir / filename
    
    with open(filepath, 'w', encoding='utf-8') as f:
        json.dump(hal_config, f, indent=2, ensure_ascii=False)
    
    return filename


# Workflow generation (for multi-station orchestration)
def generate_workflow_from_recipes(recipes: List[Dict], workflow_name: Optional[str] = None) -> Dict:
    """
    Generate a workflow from one or more recipes.
    
    Recipes can have a '_delay_after' attribute specifying delay in seconds
    before the next recipe. Wait steps are inserted between recipes.
    """
    from datetime import datetime
    
    if not recipes:
        return {'error': 'No recipes provided'}
    
    first_recipe = recipes[0]
    name = workflow_name or f"Workflow_{first_recipe.get('ProtocolName', 'Recipe')}"
    
    steps = []
    step_id = 1
    total_delay = 0
    
    for i, recipe in enumerate(recipes):
        # Add Opentrons execution step
        steps.append({
            'id': step_id,
            'type': 'opentrons_protocol',
            'name': f"Execute {recipe.get('ProtocolName', 'Recipe')}",
            'server': 'opentrons',
            'command': 'RunRecipe',
            'parameters': {
                'recipe': recipe.get('ProtocolName'),
                'hal_config': recipe.get('HALFile', 'Standard_Flex_Setup')
            },
            'timeout_minutes': 60
        })
        step_id += 1
        
        # Check for delay after this recipe
        delay_seconds = recipe.get('_delay_after', 0)
        if delay_seconds > 0 and i < len(recipes) - 1:  # Don't add delay after last recipe
            total_delay += delay_seconds
            delay_minutes = delay_seconds // 60
            delay_secs = delay_seconds % 60
            delay_str = f"{delay_minutes}min {delay_secs}s" if delay_secs else f"{delay_minutes} minutes"
            
            # Add wait step
            steps.append({
                'id': step_id,
                'type': 'wait',
                'name': f"Wait {delay_str}",
                'server': None,
                'command': 'Wait',
                'parameters': {
                    'duration_seconds': delay_seconds,
                    'message': f"Incubation period: {delay_str}"
                },
                'timeout_minutes': (delay_seconds // 60) + 5
            })
            step_id += 1
            
            # Add extension point for optional Tecan measurement during wait
            steps.append({
                'id': step_id,
                'type': 'extension_point',
                'name': f"Optional: Tecan measurement during incubation",
                'server': 'tecan',
                'command': None,
                'parameters': {},
                'description': 'Add Tecan commands here to measure during incubation'
            })
            step_id += 1
    
    # Final extension point for Tecan analysis
    steps.append({
        'id': step_id,
        'type': 'extension_point',
        'name': 'Final Analysis (optional)',
        'server': 'tecan',
        'command': None,
        'parameters': {},
        'description': 'Add final Tecan analysis commands here'
    })
    
    return {
        'name': name,
        'version': '1.0',
        'created': datetime.now().isoformat(),
        'description': f'Generated workflow with {len(recipes)} recipe(s), total wait time: {total_delay//60} min',
        'steps': steps,
        'metadata': {
            'total_recipes': len(recipes),
            'total_wait_seconds': total_delay,
            'source': 'excel_recipe_parser'
        }
    }


def save_workflow_to_library(workflow: Dict, library_path: Path) -> str:
    """Save workflow JSON to Library/Workflows."""
    workflows_dir = Path(library_path) / "Workflows"
    workflows_dir.mkdir(parents=True, exist_ok=True)
    
    filename = f"{workflow['name']}.json"
    filepath = workflows_dir / filename
    
    with open(filepath, 'w', encoding='utf-8') as f:
        json.dump(workflow, f, indent=2)
    
    return filename


# Legacy compatibility - wrapper for old API
def parse_excel_recipe(file_content: bytes, filename: str = "recipe.xlsx") -> Dict[str, Any]:
    """
    Legacy function for backward compatibility.
    Parses Excel and generates recipes with default config (no delays).
    """
    result = parse_excel_columns(file_content, filename)
    
    if 'error' in result:
        return result
    
    all_recipes = []
    
    for sheet_data in result.get('sheets', []):
        if 'error' in sheet_data:
            continue
        
        # Use default config (all reagents, no delays)
        config = generate_default_config(sheet_data)
        recipe = generate_recipe_with_config(sheet_data, config)
        
        if 'error' not in recipe:
            all_recipes.append(recipe)
    
    return {
        'filename': filename,
        'sheets': result.get('sheets', []),
        'recipes': all_recipes,
        'summary': {
            **result.get('summary', {}),
            'total_recipes': len(all_recipes)
        }
    }


# For command-line testing
if __name__ == "__main__":
    import sys
    import json
    
    if len(sys.argv) < 2:
        print("Usage: python excel_recipe_parser.py <excel_file>")
        sys.exit(1)
    
    filepath = Path(sys.argv[1])
    
    with open(filepath, 'rb') as f:
        content = f.read()
    
    result = parse_excel_columns(content, filepath.name)
    logging.basicConfig(level=logging.INFO)
    _log = logging.getLogger(__name__)
    _log.info("=== Parsed Columns ===")
    _log.info(json.dumps(result, indent=2, default=str))
    
    # Generate with custom config example
    if result.get('sheets') and 'error' not in result['sheets'][0]:
        sheet = result['sheets'][0]
        config = {
            'groups': [
                {
                    'name': 'Phase A',
                    'columns': sheet['summary']['reagents'][:2] if len(sheet['summary']['reagents']) >= 2 else sheet['summary']['reagents'],
                    'delay_after_seconds': 1800
                },
                {
                    'name': 'Phase B',
                    'columns': sheet['summary']['reagents'][2:] if len(sheet['summary']['reagents']) > 2 else [],
                    'delay_after_seconds': 0
                }
            ],
            'pipette': 'left'
        }
        
        recipe = generate_recipe_with_config(sheet, config, 'Test_Recipe')
        _log.info("=== Generated Recipe ===")
        _log.info(json.dumps(recipe, indent=2, default=str))
